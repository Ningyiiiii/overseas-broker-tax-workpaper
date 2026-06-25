"""Excel workbook builder for tax workpapers.

Generates workbooks with 11 sheets (12 for US) per the output workbook spec.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from tax_engines import (
    calculate_fifo, calculate_period_weighted_average,
    get_period_key, get_period_end_date, get_period_start_date,
    FifoResult, PwaResult,
)


# ---- Styles ----
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True)
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name="微软雅黑", size=10)
BOLD_FONT = Font(name="微软雅黑", size=10, bold=True)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
EXCEPTION_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")


# ---- PBOC FX rates (USD to HKD) ----
PBOC_FX_RATES = {
    "2021-12-31": 7.8051,
    "2022-03-31": 7.8315,
    "2022-12-31": 7.8087,
    "2023-03-31": 7.8495,
}


def _get_fx_rate(period_end_date: str) -> tuple[float, str]:
    """Get PBOC FX rate for period end date, or nearest previous."""
    target = period_end_date
    # Try exact date first
    if target in PBOC_FX_RATES:
        return PBOC_FX_RATES[target], target
    # Find nearest previous
    available = sorted(PBOC_FX_RATES.keys(), reverse=True)
    for d in available:
        if d <= target:
            return PBOC_FX_RATES[d], d
    if available:
        return PBOC_FX_RATES[available[-1]], available[-1]
    return 0.0, target


def _set_column_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _style_header_row(ws, row: int, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def _style_total_row(ws, row: int, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = BOLD_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER


def _style_data_cell(cell, is_exception=False):
    cell.font = NORMAL_FONT
    cell.border = THIN_BORDER
    if is_exception:
        cell.fill = EXCEPTION_FILL
    if isinstance(cell.value, (int, float)):
        cell.alignment = RIGHT_ALIGN
    else:
        cell.alignment = LEFT_ALIGN


# ---- Sheet builders ----

def _build_annual_total_sheet(ws, market: str, period_regime: str, cost_method: str,
                              fifo_result=None, pwa_result=None, incomes=None,
                              financing=None, period_keys: list[str] = None):
    """Build 年度合计 sheet."""
    _set_column_widths(ws, [4, 16, 14, 14, 14, 14, 14])

    regime_label = "中国自然年" if period_regime == "china_calendar_year" else "香港财年"
    method_label = "FIFO" if cost_method == "fifo" else "期间加权平均成本法"
    market_label = "港股" if market == "HK" else "美股"
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{market_label} {regime_label} {method_label} 年度合计"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER_ALIGN

    headers = ["序号", "期间", "卖出总额", "交易费用合计", "盈亏(P&L)合计", "股息合计", "融资利息合计"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _style_header_row(ws, 3, len(headers))

    result = fifo_result if cost_method == "fifo" else pwa_result
    if not result or not period_keys:
        ws.cell(row=4, column=1, value="暂无数据")
        ws.cell(row=4, column=1).font = NORMAL_FONT
        return

    # Calculate income totals by period (净额口径: 累加所有 income 含税费扣减,
    # 与"股息利息_公司行动"分页合计保持一致, 符合 skill validation_checklist 要求)
    income_by_period: dict[str, float] = {}
    if incomes:
        for inc in incomes:
            if inc.market != market:
                continue
            pk = get_period_key(inc.date, period_regime)
            if pk not in income_by_period:
                income_by_period[pk] = 0.0
            income_by_period[pk] += inc.amount

    # Calculate financing interest totals by period
    fin_by_period: dict[str, float] = {}
    if financing:
        for f in financing:
            if f.market != market:
                continue
            pk = get_period_key(f.date, period_regime)
            if pk not in fin_by_period:
                fin_by_period[pk] = 0.0
            fin_by_period[pk] += f.amount

    row = 4
    for seq, pk in enumerate(period_keys, 1):
        totals = result.period_total.get(pk, {})
        div_total = income_by_period.get(pk, 0.0)
        fin_total = fin_by_period.get(pk, 0.0)
        values = [
            seq, pk,
            totals.get("sell_total", 0),
            totals.get("fee_total", 0),
            totals.get("pnl_total", 0),
            div_total,
            fin_total,
        ]
        for i, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=i, value=v)
            _style_data_cell(cell)
        row += 1


def _build_summary_sheet(ws, market: str, period_regime: str, cost_method: str,
                         fifo_result=None, pwa_result=None, period_keys: list[str] = None):
    """Build 汇总 sheet."""
    _set_column_widths(ws, [4, 16, 14, 14, 14])

    regime_label = "中国自然年" if period_regime == "china_calendar_year" else "香港财年"
    method_label = "FIFO" if cost_method == "fifo" else "期间加权平均成本法"
    market_label = "港股" if market == "HK" else "美股"
    ws.merge_cells("A1:E1")
    ws["A1"] = f"{market_label} {regime_label} {method_label} 汇总"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER_ALIGN

    headers = ["序号", "期间", "卖出总额", "交易费用合计", "盈亏(P&L)合计"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _style_header_row(ws, 3, len(headers))

    result = fifo_result if cost_method == "fifo" else pwa_result
    if not result or not period_keys:
        ws.cell(row=4, column=1, value="暂无数据")
        ws.cell(row=4, column=1).font = NORMAL_FONT
        return

    row = 4
    for seq, pk in enumerate(period_keys, 1):
        totals = result.period_total.get(pk, {})
        values = [seq, pk, totals.get("sell_total", 0), totals.get("fee_total", 0), totals.get("pnl_total", 0)]
        for i, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=i, value=v)
            _style_data_cell(cell)
        row += 1


def _build_fifo_trade_sheet(ws, market: str, period_regime: str, period_key: str, fifo_result: FifoResult):
    """Build a FIFO period detail sheet."""
    _set_column_widths(ws, [4, 12, 10, 16, 10, 10, 10, 12, 12, 12, 10, 10, 12, 12, 12, 12, 12, 20])

    market_label = "港股" if market == "HK" else "美股"
    ws.merge_cells("A1:R1")
    ws["A1"] = f"{market_label} {period_key} FIFO 税务底稿"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER_ALIGN

    headers = ["序号", "卖出日期", "证券代码", "证券名称", "买/卖", "卖出数量", "卖出价格",
               "卖出金额", "卖出费用分摊", "买入日期", "买入价格", "买入数量", "买入金额分摊",
               "买入费用分摊", "成本基础", "交易费用合计", "盈亏(P&L)", "备注"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _style_header_row(ws, 3, len(headers))

    # Filter details for this period
    period_details = [d for d in fifo_result.details if get_period_key(d.sell_date, period_regime) == period_key]
    if not period_details:
        ws.cell(row=4, column=1, value="暂无数据")
        ws.cell(row=4, column=1).font = NORMAL_FONT
        return

    row = 4
    for seq, d in enumerate(period_details, 1):
        values = [
            seq, d.sell_date, d.code, d.name, d.side,
            d.sell_quantity, d.sell_price, d.sell_amount, d.sell_fee_allocated,
            d.buy_date, d.buy_price, d.buy_quantity, d.buy_amount_allocated,
            d.buy_fee_allocated, d.cost_basis, d.transaction_fee, d.pnl, d.remark,
        ]
        for i, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=i, value=v)
            _style_data_cell(cell, is_exception=d.is_split_summary or (d.pnl is None and d.side == "卖出"))
        row += 1

    # Period total row
    totals = fifo_result.period_total.get(period_key, {})
    ws.cell(row=row, column=1, value=None)
    ws.cell(row=row, column=2, value=f"{period_key} 合计")
    ws.cell(row=row, column=8, value=totals.get("sell_total", 0))
    ws.cell(row=row, column=9, value=totals.get("fee_total", 0))
    ws.cell(row=row, column=15, value=None)
    ws.cell(row=row, column=16, value=totals.get("fee_total", 0))
    ws.cell(row=row, column=17, value=totals.get("pnl_total", 0))
    _style_total_row(ws, row, len(headers))


def _build_pwa_trade_sheet(ws, market: str, period_regime: str, period_key: str, pwa_result: PwaResult):
    """Build a PWA period detail sheet."""
    _set_column_widths(ws, [4, 12, 10, 16, 10, 10, 10, 12, 12, 12, 12, 12, 20])

    market_label = "港股" if market == "HK" else "美股"
    ws.merge_cells("A1:M1")
    ws["A1"] = f"{market_label} {period_key} 期间加权平均成本法 税务底稿"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER_ALIGN

    headers = ["序号", "卖出日期", "证券代码", "证券名称", "买/卖", "卖出数量", "卖出价格",
               "卖出金额", "卖出费用", "加权平均成本", "成本基础", "盈亏(P&L)", "备注"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _style_header_row(ws, 3, len(headers))

    period_details = [d for d in pwa_result.details if get_period_key(d.sell_date, period_regime) == period_key]
    if not period_details:
        ws.cell(row=4, column=1, value="暂无数据")
        ws.cell(row=4, column=1).font = NORMAL_FONT
        return

    row = 4
    for seq, d in enumerate(period_details, 1):
        values = [
            seq, d.sell_date, d.code, d.name, d.side,
            d.sell_quantity, d.sell_price, d.sell_amount, d.sell_fee,
            d.weighted_avg_cost, d.cost_basis, d.pnl, d.remark,
        ]
        for i, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=i, value=v)
            _style_data_cell(cell, is_exception=(d.pnl is None))
        row += 1

    # Period total row
    totals = pwa_result.period_total.get(period_key, {})
    ws.cell(row=row, column=1, value=None)
    ws.cell(row=row, column=2, value=f"{period_key} 合计")
    ws.cell(row=row, column=8, value=totals.get("sell_total", 0))
    ws.cell(row=row, column=9, value=totals.get("fee_total", 0))
    ws.cell(row=row, column=12, value=totals.get("pnl_total", 0))
    _style_total_row(ws, row, len(headers))


def _build_income_sheet(ws, market: str, period_regime: str, incomes: list):
    """Build 股息利息_公司行动 sheet."""
    _set_column_widths(ws, [4, 12, 10, 16, 12, 10, 12, 14, 30])

    market_label = "港股" if market == "HK" else "美股"
    ws.merge_cells("A1:I1")
    ws["A1"] = f"{market_label} 股息利息_公司行动"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER_ALIGN

    headers = ["序号", "日期", "证券代码", "证券名称", "类别", "币种", "金额", "年度合计", "备注"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _style_header_row(ws, 3, len(headers))

    market_incomes = [i for i in incomes if i.market == market]
    if not market_incomes:
        ws.cell(row=4, column=1, value="暂无数据")
        ws.cell(row=4, column=1).font = NORMAL_FONT
        return

    # Group by period
    by_period: dict[str, list] = {}
    for inc in market_incomes:
        pk = get_period_key(inc.date, period_regime)
        if pk not in by_period:
            by_period[pk] = []
        by_period[pk].append(inc)

    row = 4
    for pk in sorted(by_period.keys()):
        period_total = 0.0
        seq = 1
        for inc in by_period[pk]:
            period_total += inc.amount
            values = [seq, inc.date, inc.code or "", inc.name or "", inc.category,
                      inc.currency, inc.amount, None, inc.raw_text]
            for i, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=i, value=v)
                _style_data_cell(cell)
            row += 1
            seq += 1

        # Period total row
        ws.cell(row=row, column=1, value=None)
        ws.cell(row=row, column=2, value=f"{pk} 合计")
        ws.cell(row=row, column=8, value=period_total)
        _style_total_row(ws, row, len(headers))
        row += 1


def _build_financing_sheet(ws, market: str, period_regime: str, financing: list):
    """Build 融资利息 sheet."""
    _set_column_widths(ws, [4, 12, 10, 14, 30])

    market_label = "港股" if market == "HK" else "美股"
    ws.merge_cells("A1:E1")
    ws["A1"] = f"{market_label} 融资利息"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER_ALIGN

    headers = ["序号", "日期", "币种", "金额", "备注"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _style_header_row(ws, 3, len(headers))

    market_fin = [f for f in financing if f.market == market]
    if not market_fin:
        ws.cell(row=4, column=1, value="暂无数据")
        ws.cell(row=4, column=1).font = NORMAL_FONT
        return

    by_period: dict[str, list] = {}
    for f in market_fin:
        pk = get_period_key(f.date, period_regime)
        if pk not in by_period:
            by_period[pk] = []
        by_period[pk].append(f)

    row = 4
    for pk in sorted(by_period.keys()):
        period_total = 0.0
        seq = 1
        for f in by_period[pk]:
            period_total += f.amount
            values = [seq, f.date, f.currency, f.amount, f.raw_text]
            for i, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=i, value=v)
                _style_data_cell(cell)
            row += 1
            seq += 1
        ws.cell(row=row, column=1, value=None)
        ws.cell(row=row, column=2, value=f"{pk} 合计")
        ws.cell(row=row, column=4, value=period_total)
        _style_total_row(ws, row, len(headers))
        row += 1


def _build_cost_exception_sheet(ws, market: str, fifo_result=None, pwa_result=None, cost_method: str = "fifo"):
    """Build 缺成本与异常 sheet."""
    _set_column_widths(ws, [4, 12, 10, 16, 10, 30])

    ws.merge_cells("A1:F1")
    ws["A1"] = "缺成本与异常"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER_ALIGN

    headers = ["序号", "交易日期", "证券代码", "证券名称", "数量", "异常说明"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _style_header_row(ws, 3, len(headers))

    result = fifo_result if cost_method == "fifo" else pwa_result
    if not result or not result.exceptions:
        ws.cell(row=4, column=1, value="暂无异常")
        ws.cell(row=4, column=1).font = NORMAL_FONT
        return

    row = 4
    for seq, exc in enumerate(result.exceptions, 1):
        values = [seq, exc["trade_date"], exc["code"], exc["name"], exc["quantity"], exc["reason"]]
        for i, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=i, value=v)
            _style_data_cell(cell, is_exception=True)
        row += 1


def _build_name_exception_sheet(ws, trades: list, incomes: list):
    """Build 缺名称与异常 sheet."""
    _set_column_widths(ws, [4, 10, 16, 30])

    ws.merge_cells("A1:D1")
    ws["A1"] = "缺名称与异常"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER_ALIGN

    headers = ["序号", "证券代码", "证券名称", "异常说明"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _style_header_row(ws, 3, len(headers))

    row = 4
    seen = set()
    found_any = False
    for record in list(trades) + list(incomes):
        code = getattr(record, "code", "")
        name = getattr(record, "name", "")
        if code and (not name or "\ufffd" in name):
            key = (code, name)
            if key not in seen:
                seen.add(key)
                msg = "名称缺失" if not name else "名称乱码"
                values = [row - 3, code, name, msg]
                for i, v in enumerate(values, 1):
                    cell = ws.cell(row=row, column=i, value=v)
                    _style_data_cell(cell, is_exception=True)
                row += 1
                found_any = True
    if not found_any:
        ws.cell(row=row, column=1, value="暂无异常")
        ws.cell(row=row, column=1).font = NORMAL_FONT


def _build_source_sheet(ws, statements: list, errors: list):
    """Build 解析来源与校验 sheet."""
    _set_column_widths(ws, [4, 30, 12, 10, 10, 10, 20])

    ws.merge_cells("A1:G1")
    ws["A1"] = "解析来源与校验"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER_ALIGN

    headers = ["序号", "文件名", "结单日期", "交易数", "收入数", "融资利息数", "状态"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _style_header_row(ws, 3, len(headers))

    row = 4
    error_files = {e["file"] for e in errors}
    for seq, stmt in enumerate(statements, 1):
        status = "解析失败" if stmt.source_file in error_files else "成功"
        values = [seq, stmt.source_file, stmt.statement_date or "",
                  len(stmt.trades), len(stmt.incomes), len(stmt.financing_interests), status]
        for i, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=i, value=v)
            _style_data_cell(cell)
        row += 1
    for e in errors:
        values = [row - 3, e["file"], "", "", "", "", f"失败: {e['error']}"]
        for i, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=i, value=v)
            _style_data_cell(cell, is_exception=True)
        row += 1


def _build_trade_source_sheet(ws, trades: list, market: str):
    """Build 交易来源明细 sheet."""
    _set_column_widths(ws, [4, 12, 10, 16, 8, 10, 10, 12, 12, 12, 20])

    ws.merge_cells("A1:K1")
    ws["A1"] = "交易来源明细"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER_ALIGN

    headers = ["序号", "交易日期", "证券代码", "证券名称", "买/卖", "数量", "价格", "金额", "费用", "来源文件", "原始文本"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _style_header_row(ws, 3, len(headers))

    market_trades = [t for t in trades if t.market == market]
    if not market_trades:
        ws.cell(row=4, column=1, value="暂无数据")
        ws.cell(row=4, column=1).font = NORMAL_FONT
        return

    row = 4
    for seq, t in enumerate(market_trades, 1):
        values = [seq, t.trade_date, t.code, t.name, t.side, t.quantity, t.price,
                  t.gross_amount, t.fee_total, t.source_file, t.raw_text]
        for i, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=i, value=v)
            _style_data_cell(cell)
        row += 1


def _build_security_dict_sheet(ws, trades: list, incomes: list):
    """Build 证券名称字典 sheet."""
    _set_column_widths(ws, [4, 8, 10, 20, 20])

    ws.merge_cells("A1:E1")
    ws["A1"] = "证券名称字典"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER_ALIGN

    headers = ["序号", "市场", "证券代码", "证券名称", "首次来源"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _style_header_row(ws, 3, len(headers))

    name_map: dict[tuple, tuple] = {}  # (market, code) -> (name, source_file)
    for t in trades:
        key = (t.market, t.code)
        if key not in name_map and t.name:
            name_map[key] = (t.name, t.source_file)
    for inc in incomes:
        key = (inc.market, inc.code)
        if key not in name_map and inc.name:
            name_map[key] = (inc.name, inc.source_file)

    row = 4
    for seq, ((mk, code), (name, source)) in enumerate(sorted(name_map.items()), 1):
        values = [seq, mk, code, name, source]
        for i, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=i, value=v)
            _style_data_cell(cell)
        row += 1


def _build_us_fx_info_sheet(ws, period_regime: str, period_keys: list[str]):
    """Build 美股汇率信息 sheet (US workbooks only)."""
    _set_column_widths(ws, [4, 16, 14, 14, 20])

    ws.merge_cells("A1:E1")
    ws["A1"] = "美股汇率信息"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER_ALIGN

    headers = ["序号", "期间", "期间截止日", "USD/HKD汇率", "汇率日期"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _style_header_row(ws, 3, len(headers))

    row = 4
    for seq, pk in enumerate(period_keys, 1):
        end_date = get_period_end_date(pk)
        rate, fx_date = _get_fx_rate(end_date)
        values = [seq, pk, end_date, rate, fx_date]
        for i, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=i, value=v)
            _style_data_cell(cell)
        row += 1


# ---- Main workbook builder ----

def build_workbook(output_path: Path, market: str, period_regime: str, cost_method: str,
                   all_trades: list, all_incomes: list, all_financing: list,
                   statements: list, errors: list):
    """Build a complete workbook for one market/regime/method combination."""
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Calculate results
    fifo_result = None
    pwa_result = None
    if cost_method == "fifo":
        fifo_result = calculate_fifo(all_trades, period_regime, market)
    else:
        pwa_result = calculate_period_weighted_average(all_trades, period_regime, market)

    # Determine period keys
    result = fifo_result if cost_method == "fifo" else pwa_result
    period_keys = sorted(result.period_total.keys()) if result else []

    # Also check incomes for additional periods
    for inc in all_incomes:
        if inc.market == market:
            pk = get_period_key(inc.date, period_regime)
            if pk not in period_keys:
                period_keys.append(pk)
    period_keys = sorted(set(period_keys))

    # Sheet: 年度合计
    ws = wb.create_sheet("年度合计")
    _build_annual_total_sheet(ws, market, period_regime, cost_method,
                              fifo_result, pwa_result, all_incomes,
                              all_financing, period_keys)

    # Sheet: 汇总
    ws = wb.create_sheet("汇总")
    _build_summary_sheet(ws, market, period_regime, cost_method,
                         fifo_result, pwa_result, period_keys)

    # Sheets: per period
    for pk in period_keys:
        ws = wb.create_sheet(pk)
        if cost_method == "fifo":
            _build_fifo_trade_sheet(ws, market, period_regime, pk, fifo_result)
        else:
            _build_pwa_trade_sheet(ws, market, period_regime, pk, pwa_result)

    # Sheet: 股息利息_公司行动
    ws = wb.create_sheet("股息利息_公司行动")
    _build_income_sheet(ws, market, period_regime, all_incomes)

    # Sheet: 融资利息
    ws = wb.create_sheet("融资利息")
    _build_financing_sheet(ws, market, period_regime, all_financing)

    # Sheet: 缺成本与异常
    ws = wb.create_sheet("缺成本与异常")
    _build_cost_exception_sheet(ws, market, fifo_result, pwa_result, cost_method)

    # Sheet: 缺名称与异常
    ws = wb.create_sheet("缺名称与异常")
    _build_name_exception_sheet(ws, all_trades, all_incomes)

    # Sheet: 解析来源与校验
    ws = wb.create_sheet("解析来源与校验")
    _build_source_sheet(ws, statements, errors)

    # Sheet: 交易来源明细
    ws = wb.create_sheet("交易来源明细")
    _build_trade_source_sheet(ws, all_trades, market)

    # Sheet: 证券名称字典
    ws = wb.create_sheet("证券名称字典")
    _build_security_dict_sheet(ws, all_trades, all_incomes)

    # Sheet: 美股汇率信息 (US only)
    if market == "US":
        ws = wb.create_sheet("美股汇率信息")
        _build_us_fx_info_sheet(ws, period_regime, period_keys)

    wb.save(str(output_path))
