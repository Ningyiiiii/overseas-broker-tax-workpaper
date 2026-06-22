from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from dataclasses import asdict
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


ROOT = Path.cwd()
PASSWORD = ""
RUN_STAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_ROOT = ROOT / "outputs" / f"tax_workpaper_new_rules_{RUN_STAMP}"
WORK_DIR = OUTPUT_ROOT / "_work"
RUNTIME_DIR = Path(__file__).resolve().parent
if str(RUNTIME_DIR) not in sys.path:
    sys.path.insert(0, str(RUNTIME_DIR))
import futu_tax_extract as futu


def configure_runtime(source_root: Path | str, output_dir: Path | str | None = None, password: str | None = None) -> None:
    """Configure source/output paths before running the generated workpaper flow."""

    global ROOT, PASSWORD, RUN_STAMP, OUTPUT_ROOT, WORK_DIR
    ROOT = Path(source_root).resolve()
    if password:
        PASSWORD = password
    RUN_STAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_output = Path(output_dir).resolve() if output_dir else ROOT / "outputs"
    OUTPUT_ROOT = base_output / f"tax_workpaper_new_rules_{RUN_STAMP}"
    WORK_DIR = OUTPUT_ROOT / "_work"


CN = {
    "date": "交易日期",
    "code": "股票代码",
    "name": "股票名称",
    "sell_price": "卖出价格",
    "buy_price": "买入价格",
    "qty": "交易数量",
    "sell_amount": "卖出总金额",
    "buy_amount": "买入总金额",
    "fee": "交易费用",
    "pnl": "盈亏",
    "note": "备注：买入时间",
    "currency": "币种",
    "source": "来源",
    "exception": "异常",
    "order_id": "订单编号",
    "fx_hkd": "折算汇率",
}

BASE_HEADERS = [
    CN["date"],
    CN["code"],
    CN["name"],
    CN["sell_price"],
    CN["buy_price"],
    CN["qty"],
    CN["sell_amount"],
    CN["buy_amount"],
    CN["fee"],
    CN["pnl"],
    CN["note"],
    CN["currency"],
    CN["order_id"],
    CN["source"],
    CN["exception"],
]

US_HKD_HEADERS = ["卖出总金额(HKD)", "买入总金额(HKD)", "交易费用(HKD)", "盈亏(HKD)", CN["fx_hkd"]]

SIDE_SET = {"买入", "卖出", "买入开仓", "卖出平仓", "買入", "賣出", "買入開倉", "賣出平倉"}
BUY_MARKERS = {"买入", "买入开仓", "買入", "買入開倉"}
SELL_MARKERS = {"卖出", "卖出平仓", "賣出", "賣出平倉"}
VENUES = {
    "SEHK",
    "FUTU OTC",
    "EDGX",
    "BATS",
    "MEMX",
    "XNAS",
    "XNYS",
    "NYSE",
    "NASDAQ",
    "AMEX",
    "ARCA",
    "ARCX",
    "IEX",
    "KNEM",
    "JNST",
    "OCEA",
    "CDED",
    "EPRL",
    "XBOS",
    "BATY",
    "KNLI",
    "US",
}
CURRENCIES = {"HKD", "USD", "CNH", "JPY", "SGD"}
DATE_RE = re.compile(r"^\d{4}/\d{2}/\d{2}$")
TIME_RE = re.compile(r"^\d{1,2}:\d{2}:\d{2}$")
NUM_RE = re.compile(r"^[+-]?\d[\d,]*(?:\.\d+)?$")


FX_RATES = {
    "CY2021": {"date": "2021-12-31", "rate": "7.798068"},
    "CY2022": {"date": "2022-12-30", "rate": "7.796747"},
    "CY2023": {"date": "2023-12-29", "rate": "7.815652"},
    "CY2024": {"date": "2024-12-31", "rate": "7.762516"},
    "CY2025": {"date": "2025-12-31", "rate": "7.781936"},
    "FY2021-2022": {"date": "2022-03-31", "rate": "7.827524"},
    "FY2022-2023": {"date": "2023-03-31", "rate": "7.849693"},
    "FY2023-2024": {"date": "2024-03-29", "rate": "7.826375"},
    "FY2024-2025": {"date": "2025-03-31", "rate": "7.778464"},
    "FY2025-2026": {"date": "2026-03-31", "rate": "7.836684"},
}


def d(value: Any) -> Decimal:
    return futu.d(value)


def money(value: Decimal) -> str:
    return futu.money(value)


def qty_str(value: Decimal) -> str:
    return futu.qty_str(value)


def q4(value: Decimal) -> str:
    return str(value.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP))


def clean(value: Any) -> str:
    return re.sub(r"[ \t]+", " ", str(value or "").replace("\xa0", " ").strip())


def safe_name(value: Any) -> str:
    value = clean(value)
    if not value or re.search(r"[�锟]", value):
        return ""
    return re.sub(r"\s+", "", value) if re.search(r"[\u4e00-\u9fff]", value) else value


def parse_code_name(text: str) -> tuple[str, str]:
    text = clean(text)
    m = re.match(r"([0-9A-Z.\-]{1,12})\((.*)\)$", text)
    if m:
        return m.group(1), safe_name(m.group(2))
    m = re.match(r"([0-9A-Z.\-]{1,12})\((.*)", text)
    if m:
        return m.group(1), safe_name(m.group(2))
    parts = text.split(None, 1)
    return (parts[0], safe_name(parts[1]) if len(parts) > 1 else "") if parts else ("", "")


def market_from_venue_currency(venue: str, currency: str) -> str:
    if venue in {"SEHK", "FUTU OTC"} and currency == "HKD":
        return "HK"
    if currency == "USD" or venue in (VENUES - {"SEHK"}):
        return "US"
    return "HK"


def normalize_code_for_market(code: str, market: str) -> str:
    code = clean(code)
    if market == "HK" and code.isdigit():
        return code.zfill(5)
    return code


def looks_like_code_name(value: str) -> bool:
    return bool(re.match(r"^[0-9A-Z.\-]{1,12}\(", clean(value)))


def next_non_noise_index(flat: list[tuple[int, str]], pos: int) -> int:
    while pos < len(flat) and is_modern_noise_line(flat[pos][1]):
        pos += 1
    return pos


def is_venue_at(flat: list[tuple[int, str]], pos: int) -> bool:
    if pos >= len(flat):
        return False
    value = clean(flat[pos][1])
    if value in VENUES:
        return True
    if re.match(r"^[A-Z0-9]{3,5}$", value):
        nxt = next_non_noise_index(flat, pos + 1)
        return nxt < len(flat) and flat[nxt][1] in CURRENCIES
    return False


def read_code_name_before_venue(flat: list[tuple[int, str]], pos: int) -> tuple[str, str, int]:
    parts = [flat[pos][1]]
    probe = pos + 1
    while probe < len(flat):
        line = clean(flat[probe][1])
        if is_venue_at(flat, probe) or line in SIDE_SET:
            break
        if not is_modern_noise_line(line):
            parts.append(line)
        probe += 1
    if is_venue_at(flat, probe):
        code, name = parse_code_name("".join(parts))
        return code, name, probe
    return "", "", pos


def read_code_name_before_currency(flat: list[tuple[int, str]], pos: int) -> tuple[str, str, int]:
    parts = [flat[pos][1]]
    probe = pos + 1
    while probe < len(flat):
        line = clean(flat[probe][1])
        if line in CURRENCIES or line in SIDE_SET or DATE_RE.match(line):
            break
        if not is_modern_noise_line(line):
            parts.append(line)
        probe += 1
    if probe < len(flat) and flat[probe][1] in CURRENCIES:
        code, name = parse_code_name("".join(parts))
        return code, name, probe
    return "", "", pos


def is_modern_noise_line(value: str) -> bool:
    value = clean(value)
    if not value:
        return True
    if DATE_RE.match(value) or re.match(r"^\d+/\d+$", value) or re.match(r"^20\d{2}/\d{2}$", value):
        return True
    if re.match(r"^\d{1,3}$", value):
        return True
    if re.match(r"^\d{6,}$", value):
        return True
    if value in {"/", "-", "—"}:
        return True
    noise_prefixes = (
        "交易",
        "交易-股票",
        "买卖方向",
        "買賣方向",
        "代码名称",
        "代碼名稱",
        "交易所",
        "貨幣種類",
        "货币种类",
        "日期/時間",
        "日期/时间",
        "交收日期",
        "數量",
        "数量",
        "價格",
        "价格",
        "成交金額",
        "成交金额",
        "變動金額",
        "变动金额",
        "保證金",
        "保证金",
        "製備日期",
        "制备日期",
        "客戶姓名",
        "客户姓名",
        "賬戶號碼",
        "账户号码",
    )
    return any(value.startswith(prefix) for prefix in noise_prefixes)


def parse_trade_dt(value: str) -> date | None:
    m = re.search(r"\d{4}[-/]\d{2}[-/]\d{2}", value or "")
    if not m:
        return None
    return datetime.strptime(m.group(0).replace("/", "-"), "%Y-%m-%d").date()


def period_key_for_date(dt: date, regime: str) -> str | None:
    if regime == "calendar":
        return f"CY{dt.year}" if 2021 <= dt.year <= 2026 else None
    start = dt.year if dt.month >= 4 else dt.year - 1
    return f"FY{start}-{start + 1}" if 2021 <= start <= 2025 else None


def period_ranges(regime: str) -> list[tuple[str, date, date]]:
    if regime == "calendar":
        return [(f"CY{year}", date(year, 1, 1), date(year, 12, 31)) for year in range(2021, 2027)]
    return [(f"FY{year}-{year + 1}", date(year, 4, 1), date(year + 1, 3, 31)) for year in range(2021, 2026)]


def period_keys(regime: str) -> list[str]:
    return [p[0] for p in period_ranges(regime)]


def modern_pdf_lines(path: Path) -> list[tuple[int, list[str]]]:
    reader = PdfReader(str(path))
    if reader.is_encrypted:
        reader.decrypt(PASSWORD)
    pages: list[tuple[int, list[str]]] = []
    for page_no, page in enumerate(reader.pages, start=1):
        lines = [clean(x) for x in (page.extract_text() or "").splitlines()]
        pages.append((page_no, [x for x in lines if x]))
    return pages


def parse_modern_trades(path: Path, rel: str) -> list[Any]:
    trades: list[Any] = []
    block_no = 0
    flat: list[tuple[int, str]] = [(page_no, line) for page_no, lines in modern_pdf_lines(path) for line in lines]

    def skip_noise(pos: int) -> int:
        while pos < len(flat) and is_modern_noise_line(flat[pos][1]):
            pos += 1
        return pos

    i = 0
    while i < len(flat):
        page_no, side_text = flat[i]
        if side_text not in SIDE_SET:
            i += 1
            continue
        if i + 1 < len(flat) and flat[i + 1][1] == side_text:
            i += 1
        side = "buy" if side_text in BUY_MARKERS else "sell"
        i = skip_noise(i + 1)

        text_order = ""
        name_parts: list[str] = []
        if i < len(flat) and flat[i][1] in CURRENCIES:
            # Text order B: side -> currency -> summary -> code(name) -> fills.
            text_order = "B"
            summary_currency = flat[i][1]
            i += 1
            if i + 3 >= len(flat) or not all(NUM_RE.match(flat[i + off][1]) for off in range(4)):
                continue
            summary_qty, summary_price, summary_amount, summary_change = [d(flat[i + off][1]) for off in range(4)]
            i += 4
            i = skip_noise(i)
            while i < len(flat) and not is_venue_at(flat, i) and flat[i][1] not in SIDE_SET:
                if not is_modern_noise_line(flat[i][1]):
                    name_parts.append(flat[i][1])
                i += 1
        else:
            # Text order A: side -> code(name) -> currency -> summary -> fills.
            text_order = "A"
            while i < len(flat) and flat[i][1] not in CURRENCIES and flat[i][1] not in SIDE_SET:
                if not is_modern_noise_line(flat[i][1]):
                    name_parts.append(flat[i][1])
                i += 1
            if i >= len(flat) or flat[i][1] not in CURRENCIES:
                continue
            summary_currency = flat[i][1]
            i += 1
            if i + 3 >= len(flat) or not all(NUM_RE.match(flat[i + off][1]) for off in range(4)):
                continue
            summary_qty, summary_price, summary_amount, summary_change = [d(flat[i + off][1]) for off in range(4)]
            i += 4

        if not name_parts:
            continue
        code, name = parse_code_name("".join(name_parts))
        block_no += 1
        order_id = f"{rel}#modern#{block_no}"
        fill_no = 0

        while i < len(flat):
            i = skip_noise(i)
            if i >= len(flat) or flat[i][1] in SIDE_SET:
                break

            fill_code = code
            fill_name = name
            if looks_like_code_name(flat[i][1]):
                maybe_code, maybe_name, probe = read_code_name_before_venue(flat, i)
                if probe != i:
                    fill_code = maybe_code or fill_code
                    fill_name = maybe_name or fill_name
                    i = probe

            if i + 8 >= len(flat) or not is_venue_at(flat, i) or flat[i + 1][1] not in CURRENCIES:
                break
            venue = flat[i][1]
            currency = flat[i + 1][1]
            trade_date = flat[i + 2][1]
            trade_time = flat[i + 3][1]
            settle_date = flat[i + 4][1]
            if not (DATE_RE.match(trade_date) and TIME_RE.match(trade_time) and DATE_RE.match(settle_date)):
                break
            if not all(NUM_RE.match(flat[i + off][1]) for off in range(5, 9)):
                break

            qty, price, amount, change = [d(flat[i + off][1]) for off in range(5, 9)]
            market = market_from_venue_currency(venue, currency)
            trade_code = normalize_code_for_market(fill_code, market)
            if qty > 0 and amount > 0 and change != 0:
                fill_no += 1
                fee = abs(abs(change) - amount)
                trades.append(
                    futu.Trade(
                        source_file=rel,
                        page=flat[i][0],
                        format=f"modern_state_order_{text_order}",
                        side=side,
                        order_id=f"{order_id}.{fill_no}",
                        code=trade_code,
                        name=fill_name,
                        market=market,
                        currency=currency,
                        trade_datetime=f"{trade_date} {trade_time}",
                        settle_date=settle_date,
                        quantity=qty_str(qty),
                        price=str(price),
                        amount=money(amount),
                        change_amount=money(change),
                        fee_total=money(fee),
                        fee_detail={},
                        raw=" / ".join([side_text, f"{trade_code}({fill_name})", summary_currency, qty_str(summary_qty), str(summary_price), money(summary_amount), money(summary_change)]),
                        notes="",
                    )
                )
            i += 9

        if fill_no == 0 and summary_qty > 0 and summary_amount > 0:
            fee = abs(abs(summary_change) - summary_amount)
            market = "US" if summary_currency == "USD" else "HK"
            trades.append(
                futu.Trade(
                    source_file=rel,
                    page=page_no,
                    format=f"modern_summary_fallback_order_{text_order}",
                    side=side,
                    order_id=order_id,
                    code=normalize_code_for_market(code, market),
                    name=name,
                    market=market,
                    currency=summary_currency,
                    trade_datetime="",
                    settle_date="",
                    quantity=qty_str(summary_qty),
                    price=str(summary_price),
                    amount=money(summary_amount),
                    change_amount=money(summary_change),
                    fee_total=money(fee),
                    fee_detail={},
                    raw="summary fallback",
                    notes="新格式未识别到逐笔成交明细",
                )
            )
    return trades


def parse_modern_ipo_allotments(path: Path, rel: str) -> list[Any]:
    allotments: list[dict[str, Any]] = []
    cash_by_code: dict[str, dict[str, Decimal]] = defaultdict(
        lambda: {"application": Decimal("0"), "refund": Decimal("0"), "handling": Decimal("0")}
    )

    pages = modern_pdf_lines(path)
    flat = [(page_no, line) for page_no, lines in pages for line in lines]

    def read_cash_movements() -> None:
        i = 0
        while i + 5 < len(flat):
            if not DATE_RE.match(flat[i][1]):
                i += 1
                continue
            date_value = flat[i][1]
            direction = flat[i + 1][1]
            j = i + 2
            item_type_parts: list[str] = []
            while j < len(flat) and flat[j][1] not in CURRENCIES and not DATE_RE.match(flat[j][1]):
                item_type_parts.append(flat[j][1])
                j += 1
            if j + 2 >= len(flat) or flat[j][1] not in CURRENCIES or not NUM_RE.match(flat[j + 1][1]):
                i += 1
                continue
            currency = flat[j][1]
            amount = d(flat[j + 1][1])
            k = j + 2
            desc_parts: list[str] = []
            while k < len(flat) and not DATE_RE.match(flat[k][1]) and flat[k][1] not in SIDE_SET:
                if not is_modern_noise_line(flat[k][1]):
                    desc_parts.append(flat[k][1])
                k += 1
            desc = clean(" ".join(desc_parts))
            m = re.search(r"#(?P<code>\d{3,5})", desc)
            if m and "IPO" in desc.upper():
                code = m.group("code").zfill(5)
                u = desc.upper()
                if "APPLICATION AMOUNT" in u:
                    cash_by_code[code]["application"] += abs(amount)
                elif "REFUND AMOUNT" in u:
                    cash_by_code[code]["refund"] += abs(amount)
                elif "HANDLING FEE" in u:
                    cash_by_code[code]["handling"] += abs(amount)
            i = max(k, i + 1)

    def read_asset_movements() -> None:
        i = 0
        while i + 7 < len(flat):
            if not DATE_RE.match(flat[i][1]):
                i += 1
                continue
            page_no = flat[i][0]
            date_value = flat[i][1]
            direction = flat[i + 1][1]
            j = i + 2
            type_parts: list[str] = []
            while j < len(flat) and not looks_like_code_name(flat[j][1]) and not DATE_RE.match(flat[j][1]):
                if not is_modern_noise_line(flat[j][1]):
                    type_parts.append(flat[j][1])
                j += 1
            if j + 4 >= len(flat) or not looks_like_code_name(flat[j][1]):
                i += 1
                continue
            code, name, currency_pos = read_code_name_before_currency(flat, j)
            if currency_pos == j or currency_pos + 3 >= len(flat):
                i += 1
                continue
            currency = flat[currency_pos][1]
            qty_value = flat[currency_pos + 1][1]
            amount_value = flat[currency_pos + 2][1]
            desc = flat[currency_pos + 3][1]
            joined_type = "".join(type_parts)
            if currency not in CURRENCIES or not NUM_RE.match(qty_value) or not NUM_RE.match(amount_value):
                i += 1
                continue
            raw_type_desc = (joined_type + desc).upper()
            is_ipo_allotment = "IPO" in raw_type_desc and "ALLOTMENT" in raw_type_desc
            is_reinvestment_stock = (
                ("公司行動" in joined_type or "公司行动" in joined_type)
                and ("REINV" in raw_type_desc or "SCP OPT" in raw_type_desc or "SCRIP" in raw_type_desc)
            )
            if not (is_ipo_allotment or is_reinvestment_stock):
                i += 1
                continue
            qty = abs(d(qty_value))
            amount = abs(d(amount_value))
            if qty <= 0 or amount <= 0:
                i += 1
                continue
            market = "US" if currency == "USD" else "HK"
            code = normalize_code_for_market(code, market)
            allotments.append(
                {
                    "source_file": rel,
                    "page": page_no,
                    "date": date_value,
                    "code": code,
                    "name": name,
                    "market": market,
                    "currency": currency,
                    "quantity": qty,
                    "amount": amount,
                    "format": "modern_ipo_allotment" if is_ipo_allotment else "modern_reinvestment_stock",
                    "raw": " / ".join([date_value, direction, joined_type, f"{code}({name})", currency, qty_value, amount_value, desc]),
                }
            )
            i = currency_pos + 4

    def read_inline_ipo_allotments() -> None:
        for i, (page_no, line) in enumerate(flat):
            if "IPO Allotment Qty" not in line:
                continue
            window_start = max(0, i - 16)
            window = [flat[x][1] for x in range(window_start, i + 1)]
            joined = " ".join(window)
            m_hash = re.search(r"#(?P<code>\d{3,5})", joined)
            if not m_hash:
                continue
            code = m_hash.group("code").zfill(5)
            code_pos = None
            parsed_name = ""
            for pos in range(i - 1, window_start - 1, -1):
                token = clean(flat[pos][1])
                if re.match(rf"^0*{int(code)}\(", token):
                    code_pos = pos
                    code_token, parsed_name = read_code_name_before_venue(flat, pos)[:2]
                    if code_token:
                        code = normalize_code_for_market(code_token, "HK")
                    break
            if code_pos is None:
                continue
            currency_pos = None
            for pos in range(code_pos + 1, min(i, code_pos + 8)):
                if flat[pos][1] in CURRENCIES:
                    currency_pos = pos
                    break
            if currency_pos is None or currency_pos + 2 >= len(flat):
                continue
            qty_value = flat[currency_pos + 1][1]
            amount_value = flat[currency_pos + 2][1]
            if not (NUM_RE.match(qty_value) and NUM_RE.match(amount_value)):
                continue
            date_value = ""
            for pos in range(code_pos, window_start - 1, -1):
                if DATE_RE.match(flat[pos][1]):
                    date_value = flat[pos][1]
                    break
            if not date_value:
                continue
            qty = abs(d(qty_value))
            amount = abs(d(amount_value))
            if qty <= 0 or amount <= 0:
                continue
            allotments.append(
                {
                    "source_file": rel,
                    "page": page_no,
                    "date": date_value,
                    "code": code,
                    "name": parsed_name,
                    "market": "US" if flat[currency_pos][1] == "USD" else "HK",
                    "currency": flat[currency_pos][1],
                    "quantity": qty,
                    "amount": amount,
                    "raw": joined,
                }
            )

    read_cash_movements()
    read_asset_movements()

    trades: list[Any] = []
    seen: set[tuple[str, str, str, str]] = set()
    for item in allotments:
        key = (item["date"], item["code"], qty_str(item["quantity"]), money(item["amount"]))
        if key in seen:
            continue
        seen.add(key)
        cash = cash_by_code.get(item["code"], {})
        application = cash.get("application", Decimal("0"))
        refund = cash.get("refund", Decimal("0"))
        handling = cash.get("handling", Decimal("0"))
        cash_cost = max(Decimal("0"), application - refund)
        ipo_fee = max(Decimal("0"), cash_cost - item["amount"]) + handling
        trades.append(
            futu.Trade(
                source_file=item["source_file"],
                page=item["page"],
                format=item.get("format", "modern_ipo_allotment"),
                side="buy",
                order_id=f"{rel}#ipo#{item['date']}#{item['code']}",
                code=item["code"],
                name=item["name"],
                market=item["market"],
                currency=item["currency"],
                trade_datetime=f"{item['date']} 00:00:00",
                settle_date=item["date"],
                quantity=qty_str(item["quantity"]),
                price=q4(item["amount"] / item["quantity"]),
                amount=money(item["amount"]),
                change_amount=money(-item["amount"]),
                fee_total=money(ipo_fee),
                fee_detail={"IPO费用": money(ipo_fee)} if ipo_fee else {},
                raw=item["raw"],
                notes=(
                    f"IPO allotment cost basis; application={money(application)} refund={money(refund)} handling={money(handling)}"
                    if item.get("format") != "modern_reinvestment_stock"
                    else "Stock reinvestment/company-action cost basis"
                ),
            )
        )
    return trades


def parse_global_old_ipo_allotments(files: list[Path]) -> list[Any]:
    allotments: list[dict[str, Any]] = []
    cash_by_code: dict[str, dict[str, Decimal]] = defaultdict(
        lambda: {"application": Decimal("0"), "refund": Decimal("0"), "handling": Decimal("0")}
    )

    def read_amount(cells: list[str]) -> Decimal | None:
        for cell in cells:
            if NUM_RE.match(cell):
                try:
                    return d(cell)
                except Exception:
                    continue
        return None

    def read_cash_row(cells: list[str]) -> None:
        joined = " ".join(c for c in cells if c)
        if "IPO" not in joined.upper() or "#" not in joined:
            return
        m = re.search(r"#(?P<code>\d{3,5})", joined)
        if not m:
            return
        amount = read_amount(cells)
        if amount is None:
            return
        code = m.group("code").zfill(5)
        u = joined.upper()
        if "APPLICATION AMOUNT" in u:
            cash_by_code[code]["application"] += abs(amount)
        elif "REFUND AMOUNT" in u:
            cash_by_code[code]["refund"] += abs(amount)
        elif "HANDLING FEE" in u:
            cash_by_code[code]["handling"] += abs(amount)

    def read_allotment_row(cells: list[str], rel: str, page_no: int) -> None:
        joined = " ".join(c for c in cells if c)
        if "IPO" not in joined.upper() or "ALLOTMENT" not in joined.upper():
            return
        code = ""
        name = ""
        for cell in cells:
            parsed_code, parsed_name = parse_code_name(cell)
            if re.match(r"^\d{5}$", parsed_code):
                code = parsed_code
                name = parsed_name
                break
        m = re.search(r"#(?P<code>\d{3,5})", joined)
        if not code and m:
            code = m.group("code").zfill(5)
        if not code:
            return
        qty = None
        for cell in cells:
            if re.match(r"^[+]\d[\d,]*(?:\.\d+)?$", cell):
                qty = abs(d(cell))
                break
        if qty is None or qty <= 0:
            return
        date_value = ""
        for cell in cells:
            if DATE_RE.match(cell):
                date_value = cell
                break
        if not date_value:
            return
        allotments.append(
            {
                "source_file": rel,
                "page": page_no,
                "date": date_value,
                "code": code,
                "name": name,
                "quantity": qty,
                "raw": joined,
            }
        )

    for path in files:
        rel = str(path.relative_to(ROOT))
        if "1001231828219038" in path.name:
            continue
        try:
            with pdfplumber.open(path, password=PASSWORD) as pdf:
                for page_no, page in enumerate(pdf.pages, start=1):
                    for table in page.extract_tables() or []:
                        for row in table:
                            cells = [clean(c) for c in row if clean(c)]
                            if not cells:
                                continue
                            read_cash_row(cells)
                            read_allotment_row(cells, rel, page_no)
        except Exception:
            continue

    trades: list[Any] = []
    seen: set[tuple[str, str, str]] = set()
    for item in allotments:
        key = (item["date"], item["code"], qty_str(item["quantity"]))
        if key in seen:
            continue
        seen.add(key)
        cash = cash_by_code.get(item["code"], {})
        application = cash.get("application", Decimal("0"))
        refund = cash.get("refund", Decimal("0"))
        handling = cash.get("handling", Decimal("0"))
        amount = max(Decimal("0"), application - refund)
        if item["quantity"] <= 0 or amount <= 0:
            continue
        trades.append(
            futu.Trade(
                source_file=item["source_file"],
                page=item["page"],
                format="global_old_ipo_allotment",
                side="buy",
                order_id=f"global-old-ipo#{item['date']}#{item['code']}",
                code=item["code"],
                name=item["name"],
                market="HK",
                currency="HKD",
                trade_datetime=f"{item['date']} 00:00:00",
                settle_date=item["date"],
                quantity=qty_str(item["quantity"]),
                price=q4(amount / item["quantity"]),
                amount=money(amount),
                change_amount=money(-amount),
                fee_total=money(handling),
                fee_detail={"IPO handling": money(handling)} if handling else {},
                raw=item["raw"],
                notes=f"Global old IPO allotment cost basis; application={money(application)} refund={money(refund)} handling={money(handling)}",
            )
        )
    return trades


def parse_modern_cash(path: Path, rel: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    income: list[dict[str, Any]] = []
    financing: list[dict[str, Any]] = []
    for page_no, lines in modern_pdf_lines(path):
        i = 0
        while i < len(lines):
            if not DATE_RE.match(lines[i]) or i + 4 >= len(lines):
                i += 1
                continue
            date_value, direction, item_type, currency, amount_value = lines[i : i + 5]
            if direction not in {"增加", "減少", "减少"} or currency not in CURRENCIES or not NUM_RE.match(amount_value):
                i += 1
                continue
            j = i + 5
            desc_parts: list[str] = []
            while j < len(lines) and not DATE_RE.match(lines[j]) and lines[j] not in SIDE_SET:
                if not lines[j].startswith(("综合账户", "綜合帳戶", "制备日期", "製備日期")):
                    desc_parts.append(lines[j])
                j += 1
            desc = clean(" ".join(desc_parts))
            category = classify_cash(desc, item_type)
            if category:
                item = {
                    "date": date_value.replace("/", "-"),
                    "category": category,
                    "currency": currency,
                    "amount": money(d(amount_value)),
                    "description": desc,
                    "source_file": rel,
                    "page": page_no,
                    "code": extract_desc_code(desc),
                    "name": "",
                    "market": "US" if currency == "USD" else "HK",
                }
                if category == "融资利息":
                    financing.append(item)
                else:
                    income.append(item)
            i = max(j, i + 1)
    return dedupe_items(income), dedupe_items(financing)


def classify_cash(desc: str, item_type: str) -> str:
    u = desc.upper()
    if "ACCOUNT UPGRADE" in u or "FUND SUBSCRIPTION" in u or "FUND REDEMPTION" in u:
        return ""
    if "INTEREST FOR MONTH" in u or re.search(r"\bMARGIN\b|\bFINANC", u):
        return "融资利息"
    if "SCRIP CHARGE" in u or "HANDLING CHARGE" in u or "ADR FEE" in u:
        return "股息/分派相关费用"
    if "F/D-" in u or "I/D-" in u or "DIVIDEND" in u or "DISTRIBUTION" in u:
        return "股息/分派"
    if "公司行動" in item_type or "公司行动" in item_type:
        return "其他公司行动"
    return ""


def extract_desc_code(desc: str) -> str:
    m = re.search(r"<(?:SEHK|NYSE|NASDAQ|AMEX|XNAS|XNYS|EDGX|BATS|MEMX)\s+([0-9A-Z.\-]{1,12})\b", desc)
    if m:
        return m.group(1)
    m = re.match(r"([A-Z][A-Z0-9.\-]{0,9})\b", desc)
    return m.group(1) if m else ""


def dedupe_items(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen = set()
    out = []
    for item in items:
        key = (item["date"], item["currency"], item["amount"], item["description"], item["source_file"], item["page"])
        if key not in seen:
            seen.add(key)
            out.append(item)
    return out


def parse_us_code_name(value: str) -> tuple[str, str]:
    m = re.match(r"([A-Z][A-Z0-9.\-]{0,11})\((.*?)\)$", clean(value))
    return (m.group(1), safe_name(m.group(2))) if m else parse_code_name(value)


def parse_old_us(pdf: pdfplumber.PDF, rel: str) -> tuple[list[Any], list[dict[str, Any]], list[dict[str, Any]]]:
    trades: list[Any] = []
    income: list[dict[str, Any]] = []
    financing: list[dict[str, Any]] = []
    for page_no, page in enumerate(pdf.pages, start=1):
        for table in page.extract_tables() or []:
            for row in table:
                cells = [clean(c) for c in row]
                if len(cells) >= 10 and re.match(r"^\d{7,}$", cells[2] or "") and re.match(r"^\d{4}/\d{2}/\d{2}$", cells[4] or ""):
                    try:
                        qty, price, amount, change = d(cells[6]), d(cells[7]), d(cells[8]), d(cells[9])
                    except Exception:
                        continue
                    if qty <= 0 or amount <= 0 or change == 0:
                        continue
                    code, name = parse_us_code_name(cells[3])
                    fee = abs(abs(change) - amount)
                    trades.append(
                        futu.Trade(
                            source_file=rel,
                            page=page_no,
                            format="us_table",
                            side="buy" if change < 0 else "sell",
                            order_id=cells[2],
                            code=code,
                            name=name,
                            market="US",
                            currency="USD",
                            trade_datetime=f"{cells[4]} 00:00:00",
                            settle_date="",
                            quantity=qty_str(qty),
                            price=str(price),
                            amount=money(amount),
                            change_amount=money(change),
                            fee_total=money(fee),
                            fee_detail={},
                            raw=" | ".join(cells),
                            notes="",
                        )
                    )
                if len(cells) >= 6 and re.match(r"^\d{4}/\d{2}/\d{2}$", cells[1] or "") and re.match(r"^[+-]?\d", cells[2] or ""):
                    desc = cells[-1]
                    category = classify_cash(desc, cells[0])
                    if not category and not re.search(r"dividend|interest|adr fee", desc, re.I):
                        continue
                    item = {
                        "date": cells[1].replace("/", "-"),
                        "category": category or "股息/分派",
                        "currency": "USD",
                        "amount": money(d(cells[2])),
                        "description": desc,
                        "source_file": rel,
                        "page": page_no,
                        "code": extract_desc_code(desc),
                        "name": "",
                        "market": "US",
                    }
                    if item["category"] == "融资利息":
                        financing.append(item)
                    else:
                        income.append(item)
    return trades, dedupe_items(income), dedupe_items(financing)


def parse_old_hk(pdf: pdfplumber.PDF, rel: str) -> tuple[list[Any], list[dict[str, Any]], list[dict[str, Any]]]:
    trades = futu.parse_new_text(pdf, rel)
    if not trades or "1001231828219038" not in rel:
        try:
            old = futu.parse_old_tables(pdf, rel)
            if old:
                trades = old if not trades else trades
        except Exception:
            pass
    try:
        trades.extend(futu.extract_ipo_allotment_trades(pdf, rel))
    except Exception:
        pass
    try:
        income, financing = futu.extract_cash_items_v2(pdf, rel)
    except Exception:
        income, financing = [], []
    for trade in trades:
        trade.market = "HK"
        trade.currency = trade.currency or "HKD"
        trade.name = safe_name(trade.name)
    for item in income:
        item["market"] = "HK"
        item["date"] = clean(item.get("date", "")).replace("/", "-")
        item["category"] = classify_cash(item.get("description", ""), item.get("category", "")) or "股息/分派"
        item["amount"] = money(d(item.get("amount", "0")))
        item.setdefault("code", extract_desc_code(item.get("description", "")))
        item.setdefault("name", "")
    for item in financing:
        item["market"] = "HK"
        item["date"] = clean(item.get("date", "")).replace("/", "-")
        item["category"] = "融资利息"
        item["amount"] = money(d(item.get("amount", "0")))
        item.setdefault("code", "")
        item.setdefault("name", "")
    return trades, dedupe_items(income), dedupe_items(financing)


def parse_sources() -> tuple[list[Any], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    trades: list[Any] = []
    income: list[dict[str, Any]] = []
    financing: list[dict[str, Any]] = []
    summaries: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    files = sorted(ROOT.rglob("*.pdf"))
    for idx, path in enumerate(files, start=1):
        rel = str(path.relative_to(ROOT))
        print(f"[{idx}/{len(files)}] {rel}", flush=True)
        before = len(trades)
        try:
            if "1001231828219038" in path.name:
                t = parse_modern_trades(path, rel)
                t.extend(parse_modern_ipo_allotments(path, rel))
                inc, fin = parse_modern_cash(path, rel)
            else:
                with pdfplumber.open(path, password=PASSWORD) as pdf:
                    if "美股" in rel:
                        t, inc, fin = parse_old_us(pdf, rel)
                    else:
                        t, inc, fin = parse_old_hk(pdf, rel)
            trades.extend(t)
            income.extend(inc)
            financing.extend(fin)
            summaries.append({"file": rel, "trades": len(t), "income_items": len(inc), "financing_items": len(fin), "error": ""})
        except Exception as exc:
            errors.append({"file": rel, "error": repr(exc)})
            summaries.append({"file": rel, "trades": len(trades) - before, "income_items": 0, "financing_items": 0, "error": repr(exc)})
    old_ipo_trades = parse_global_old_ipo_allotments(files)
    if old_ipo_trades:
        trades.extend(old_ipo_trades)
        summaries.append(
            {
                "file": "__global_old_ipo_allotments__",
                "trades": len(old_ipo_trades),
                "income_items": 0,
                "financing_items": 0,
                "error": "",
            }
        )
    return trades, income, financing, summaries, errors


def backfill_names(trades: list[Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    best: dict[tuple[str, str], str] = {}
    for trade in trades:
        trade.name = safe_name(trade.name)
        key = (trade.market or "", trade.code)
        if trade.name and len(trade.name) > len(best.get(key, "")):
            best[key] = trade.name
    name_exceptions: list[dict[str, Any]] = []
    seen = set()
    for trade in trades:
        key = (trade.market or "", trade.code)
        if best.get(key):
            trade.name = best[key]
        elif key not in seen:
            seen.add(key)
            name_exceptions.append({"market": trade.market, "code": trade.code, "exception": "缺股票名称或来源名称疑似乱码", "source_file": trade.source_file})
            trade.name = ""
    master = [{"market": k[0], "code": k[1], "name": v} for k, v in sorted(best.items())]
    return master, name_exceptions


def normalized_trades(trades: list[Any]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for t in trades:
        dt = parse_trade_dt(t.trade_datetime)
        if not dt:
            continue
        venue_market = t.market
        market = "HK" if venue_market == "SEHK" else venue_market
        if t.currency == "USD":
            market = "US"
        if market not in {"HK", "US"}:
            market = "HK"
        out.append(
            {
                "side": t.side,
                "code": t.code,
                "name": safe_name(t.name),
                "market": market,
                "currency": t.currency or ("USD" if market == "US" else "HKD"),
                "trade_datetime": t.trade_datetime.replace("/", "-"),
                "trade_date": dt.isoformat(),
                "settle_date": clean(t.settle_date).replace("/", "-"),
                "quantity": qty_str(d(t.quantity)),
                "price": q4(d(t.price)),
                "amount": money(d(t.amount)),
                "fee_total": money(d(t.fee_total)),
                "change_amount": money(d(t.change_amount)),
                "order_id": t.order_id.rsplit(".", 1)[0],
                "fill_id": t.order_id,
                "source_file": t.source_file,
                "source_pages": str(t.page),
                "format": t.format,
                "notes": t.notes,
            }
        )
    return sorted(out, key=lambda x: (x["trade_date"], x["trade_datetime"], x["code"], x["side"], x["fill_id"]))


def row_from_sale(t: dict[str, Any], buy_price: str, qty: Decimal, sell_amount: Decimal, buy_amount: Any, fee: Decimal, pnl: Any, note: str, source: str, exception: str = "") -> dict[str, Any]:
    return {
        CN["date"]: t["trade_date"],
        CN["code"]: t["code"],
        CN["name"]: t["name"],
        CN["sell_price"]: dec_or_blank(t["price"]),
        CN["buy_price"]: dec_or_blank(buy_price),
        CN["qty"]: dec_or_blank(qty_str(qty)),
        CN["sell_amount"]: dec_or_blank(money(sell_amount)),
        CN["buy_amount"]: "" if buy_amount == "" else dec_or_blank(money(buy_amount)),
        CN["fee"]: dec_or_blank(money(fee)),
        CN["pnl"]: "" if pnl == "" else dec_or_blank(money(pnl)),
        CN["note"]: note,
        CN["currency"]: t["currency"],
        CN["order_id"]: t.get("order_id", ""),
        CN["source"]: source,
        CN["exception"]: exception,
    }


def dec_or_blank(value: Any) -> Decimal | str:
    if value == "" or value is None:
        return ""
    return d(value)


def calculate_fifo(trades: list[dict[str, Any]], regime: str) -> tuple[dict[str, list[dict[str, Any]]], list[dict[str, Any]]]:
    queues: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    exceptions: list[dict[str, Any]] = []
    for t in trades:
        dt = parse_trade_dt(t["trade_date"])
        if not dt:
            exceptions.append({"type": "missing trade date", **t})
            continue
        key = (t["code"], t["currency"])
        qty, amount, fee = d(t["quantity"]), d(t["amount"]), d(t["fee_total"])
        if t["side"] == "buy":
            queues[key].append({**t, "remaining_qty": qty, "remaining_amount": amount, "remaining_fee": fee})
            continue
        if t["side"] != "sell":
            continue
        period = period_key_for_date(dt, regime)
        sell_remaining, sell_amount_remaining, sell_fee_remaining = qty, amount, fee
        while sell_remaining > 0:
            while queues[key] and queues[key][0]["remaining_qty"] <= Decimal("0.00000001"):
                queues[key].pop(0)
            if not queues[key]:
                row = row_from_sale(t, "", sell_remaining, sell_amount_remaining, "", sell_fee_remaining, "", "缺买入成本", t["source_file"], "缺买入成本")
                if period:
                    rows[period].append(row)
                    exceptions.append({"type": "missing cost basis", **row})
                break
            lot = queues[key][0]
            use_qty = min(sell_remaining, lot["remaining_qty"])
            buy_ratio = use_qty / lot["remaining_qty"]
            sell_ratio = use_qty / sell_remaining
            buy_amount_part = (lot["remaining_amount"] * buy_ratio).quantize(Decimal("0.00000001"))
            buy_fee_part = (lot["remaining_fee"] * buy_ratio).quantize(Decimal("0.00000001"))
            sell_amount_part = (sell_amount_remaining * sell_ratio).quantize(Decimal("0.00000001"))
            sell_fee_part = (sell_fee_remaining * sell_ratio).quantize(Decimal("0.00000001"))
            total_fee = buy_fee_part + sell_fee_part
            pnl = sell_amount_part - buy_amount_part - total_fee
            buy_price = buy_amount_part / use_qty if use_qty else Decimal("0")
            row = row_from_sale(t, q4(buy_price), use_qty, sell_amount_part, buy_amount_part, total_fee, pnl, lot["trade_date"], f"sell:{t['source_file']} buy:{lot['source_file']}", t.get("notes", ""))
            if period:
                rows[period].append(row)
            lot["remaining_qty"] -= use_qty
            lot["remaining_amount"] -= buy_amount_part
            lot["remaining_fee"] -= buy_fee_part
            sell_remaining -= use_qty
            sell_amount_remaining -= sell_amount_part
            sell_fee_remaining -= sell_fee_part
    return rows, exceptions


def reduce_opening(pos: dict[str, Any], sell_qty: Decimal) -> None:
    if sell_qty <= 0 or pos["qty"] <= 0:
        return
    use_qty = min(sell_qty, pos["qty"])
    ratio = use_qty / pos["qty"]
    pos["cost"] -= (pos["cost"] * ratio).quantize(Decimal("0.00000001"))
    pos["qty"] -= use_qty
    if pos["qty"] <= Decimal("0.00000001"):
        pos["qty"], pos["cost"] = Decimal("0"), Decimal("0")


def build_opening(trades: list[dict[str, Any]], start: date) -> dict[tuple[str, str], dict[str, Any]]:
    positions: dict[tuple[str, str], dict[str, Any]] = defaultdict(lambda: {"qty": Decimal("0"), "cost": Decimal("0"), "first_buy_date": ""})
    for t in trades:
        dt = parse_trade_dt(t["trade_date"])
        if not dt or dt >= start:
            continue
        key = (t["code"], t["currency"])
        pos = positions[key]
        if t["side"] == "buy":
            if pos["qty"] <= 0:
                pos["first_buy_date"] = t["trade_date"]
            pos["qty"] += d(t["quantity"])
            pos["cost"] += d(t["amount"]) + d(t["fee_total"])
        elif t["side"] == "sell":
            reduce_opening(pos, d(t["quantity"]))
    return positions


def calculate_period_weighted(trades: list[dict[str, Any]], regime: str) -> tuple[dict[str, list[dict[str, Any]]], list[dict[str, Any]]]:
    rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    exceptions: list[dict[str, Any]] = []
    ranges = period_ranges(regime)
    positions = build_opening(trades, ranges[0][1])
    for period, start, end in ranges:
        period_trades = [t for t in trades if (dt := parse_trade_dt(t["trade_date"])) and start <= dt <= end]
        by_key: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
        for t in period_trades:
            by_key[(t["code"], t["currency"])].append(t)
        for key in sorted(set(positions) | set(by_key)):
            stock_trades = sorted(by_key.get(key, []), key=lambda x: (x["trade_date"], x["trade_datetime"], x["side"], x["fill_id"]))
            opening = positions[key]
            opening_qty, opening_cost = opening["qty"], opening["cost"]
            buys = [t for t in stock_trades if t["side"] == "buy"]
            sells = [t for t in stock_trades if t["side"] == "sell"]
            buy_qty = sum((d(t["quantity"]) for t in buys), Decimal("0"))
            buy_cost = sum((d(t["amount"]) + d(t["fee_total"]) for t in buys), Decimal("0"))
            available_qty, available_cost = opening_qty + buy_qty, opening_cost + buy_cost
            avg_cost = available_cost / available_qty if available_qty > 0 else Decimal("0")
            first_buy_date = opening.get("first_buy_date", "") or min((t["trade_date"] for t in buys), default="")
            sold_qty = Decimal("0")
            for t in sells:
                sell_qty, sell_amount, sell_fee = d(t["quantity"]), d(t["amount"]), d(t["fee_total"])
                if available_qty <= 0 or sold_qty >= available_qty:
                    row = row_from_sale(t, "", sell_qty, sell_amount, "", sell_fee, "", "缺买入成本", t["source_file"], "缺买入成本")
                    rows[period].append(row)
                    exceptions.append({"type": "missing cost basis", **row})
                    continue
                use_qty = min(sell_qty, available_qty - sold_qty)
                sell_ratio = use_qty / sell_qty if sell_qty else Decimal("0")
                sell_amount_part = (sell_amount * sell_ratio).quantize(Decimal("0.00000001"))
                sell_fee_part = (sell_fee * sell_ratio).quantize(Decimal("0.00000001"))
                buy_amount_part = (avg_cost * use_qty).quantize(Decimal("0.00000001"))
                pnl = sell_amount_part - buy_amount_part - sell_fee_part
                note = f"期间加权平均成本法；{period}；期初股数{qty_str(opening_qty)}；本期买入股数{qty_str(buy_qty)}；持仓起始买入日期{first_buy_date}"
                rows[period].append(row_from_sale(t, q4(avg_cost), use_qty, sell_amount_part, buy_amount_part, sell_fee_part, pnl, note, f"sell:{t['source_file']} period_weighted_average_pool", t.get("notes", "")))
                sold_qty += use_qty
                missing_qty = sell_qty - use_qty
                if missing_qty > 0:
                    missing_ratio = missing_qty / sell_qty
                    missing_row = row_from_sale(t, "", missing_qty, (sell_amount * missing_ratio).quantize(Decimal("0.00000001")), "", (sell_fee * missing_ratio).quantize(Decimal("0.00000001")), "", "缺买入成本", t["source_file"], "缺买入成本")
                    rows[period].append(missing_row)
                    exceptions.append({"type": "missing cost basis", **missing_row})
            closing_qty = available_qty - sold_qty
            closing_cost = (avg_cost * closing_qty).quantize(Decimal("0.00000001")) if closing_qty > 0 else Decimal("0")
            positions[key] = {"qty": closing_qty, "cost": closing_cost, "first_buy_date": first_buy_date if closing_qty > 0 else ""}
    return rows, exceptions


def add_fx_to_rows(rows_by_period: dict[str, list[dict[str, Any]]], market: str) -> list[dict[str, Any]]:
    fx_exceptions: list[dict[str, Any]] = []
    if market != "US":
        return fx_exceptions
    for period, rows in rows_by_period.items():
        fx = FX_RATES.get(period)
        if not fx:
            for row in rows:
                row.update({h: "" for h in US_HKD_HEADERS})
                row[CN["exception"]] = "; ".join(x for x in [str(row.get(CN["exception"], "")), "缺少年末USD/HKD汇率，未折算HKD"] if x)
            if rows:
                fx_exceptions.append({"period": period, "exception": "缺少年末USD/HKD汇率，未折算HKD"})
            continue
        rate = Decimal(str(fx["rate"]))
        label = f"USD/HKD {fx['rate']} @ {fx['date']}"
        for row in rows:
            row["卖出总金额(HKD)"] = "" if row[CN["sell_amount"]] == "" else dec_or_blank(money(d(row[CN["sell_amount"]]) * rate))
            row["买入总金额(HKD)"] = "" if row[CN["buy_amount"]] == "" else dec_or_blank(money(d(row[CN["buy_amount"]]) * rate))
            row["交易费用(HKD)"] = "" if row[CN["fee"]] == "" else dec_or_blank(money(d(row[CN["fee"]]) * rate))
            row["盈亏(HKD)"] = "" if row[CN["pnl"]] == "" else dec_or_blank(money(d(row[CN["pnl"]]) * rate))
            row[CN["fx_hkd"]] = label
    return fx_exceptions


def income_period_sums(items: list[dict[str, Any]], regime: str, market: str) -> dict[str, dict[str, Decimal]]:
    sums: dict[str, dict[str, Decimal]] = defaultdict(lambda: defaultdict(Decimal))
    for item in items:
        if item.get("market") != market:
            continue
        dt = parse_trade_dt(item.get("date", ""))
        if not dt:
            continue
        period = period_key_for_date(dt, regime)
        if period:
            sums[period][item.get("currency", "") or ("USD" if market == "US" else "HKD")] += d(item.get("amount", "0"))
    return sums


def insert_order_summary(rows: list[dict[str, Any]], headers: list[str]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    i = 0
    while i < len(rows):
        order_id = rows[i].get(CN["order_id"], "")
        group = [rows[i]]
        j = i + 1
        while j < len(rows) and order_id and rows[j].get(CN["order_id"], "") == order_id:
            group.append(rows[j])
            j += 1
        if len(group) > 1:
            first = group[0]
            summary = {h: "" for h in headers}
            summary.update(
                {
                    CN["date"]: first.get(CN["date"], ""),
                    CN["code"]: first.get(CN["code"], ""),
                    CN["name"]: first.get(CN["name"], ""),
                    CN["sell_price"]: first.get(CN["sell_price"], ""),
                    CN["qty"]: sum_dec(group, CN["qty"]),
                    CN["sell_amount"]: sum_dec(group, CN["sell_amount"]),
                    CN["fee"]: sum_dec(group, CN["fee"]),
                    CN["note"]: "订单汇总（不参与盈亏计算）",
                    CN["currency"]: first.get(CN["currency"], ""),
                    CN["order_id"]: order_id,
                    CN["source"]: "订单分拆成交汇总",
                }
            )
            out.append(summary)
        out.extend(group)
        i = j
    return out


def sum_dec(rows: list[dict[str, Any]], key: str) -> Decimal | str:
    values = [d(r[key]) for r in rows if r.get(key) != ""]
    return sum(values, Decimal("0")) if values else ""


def append_period_totals(rows: list[dict[str, Any]], headers: list[str], income_sum: dict[str, Decimal], financing_sum: dict[str, Decimal]) -> list[dict[str, Any]]:
    rows = list(rows)
    total = {h: "" for h in headers}
    total[CN["date"]] = "年度合计"
    total[CN["sell_amount"]] = sum_dec(rows, CN["sell_amount"])
    total[CN["fee"]] = sum_dec(rows, CN["fee"])
    total[CN["pnl"]] = sum_dec(rows, CN["pnl"])
    rows.append(total)
    for currency, value in sorted(income_sum.items()):
        row = {h: "" for h in headers}
        row[CN["date"]] = "股息/分派合计"
        row[CN["currency"]] = currency
        row["股息/分派合计"] = value
        row["股息/分派按年合计"] = value
        rows.append(row)
    for currency, value in sorted(financing_sum.items()):
        row = {h: "" for h in headers}
        row[CN["date"]] = "融资利息合计"
        row[CN["currency"]] = currency
        row["融资利息合计"] = value
        rows.append(row)
    return rows


def write_sheet(ws, rows: list[dict[str, Any]], headers: list[str]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    style_sheet(ws)


def style_sheet(ws) -> None:
    fill = PatternFill("solid", fgColor="FFFF00")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    for col in range(1, ws.max_column + 1):
        max_len = 8
        for cell in ws.iter_rows(min_col=col, max_col=col, values_only=True):
            value = "" if cell[0] is None else str(cell[0])
            max_len = max(max_len, min(45, len(value) + 2))
        ws.column_dimensions[get_column_letter(col)].width = max_len


def write_workbook(path: Path, payload: dict[str, Any]) -> None:
    market = payload["market"]
    headers = list(BASE_HEADERS)
    if market == "US":
        headers.extend(US_HKD_HEADERS)
    headers.extend(["股息/分派合计", "股息/分派按年合计", "融资利息合计"])
    wb = Workbook()
    ws = wb.active
    ws.title = "年度合计"
    summary_rows = []
    for period in payload["period_keys"]:
        rows = payload["rows"].get(period, [])
        summary_rows.append(
            {
                "期间": period,
                "记录数": len(rows),
                "缺成本记录数": sum(1 for r in rows if r.get(CN["exception"]) == "缺买入成本"),
                "卖出总金额合计": sum_dec(rows, CN["sell_amount"]),
                "交易费用合计": sum_dec(rows, CN["fee"]),
                "盈亏合计": sum_dec(rows, CN["pnl"]),
                "股息/分派合计": "; ".join(f"{ccy} {money(val)}" for ccy, val in sorted(payload["income_sums"].get(period, {}).items())),
                "股息/分派按年合计": "; ".join(f"{ccy} {money(val)}" for ccy, val in sorted(payload["income_sums"].get(period, {}).items())),
                "融资利息合计": "; ".join(f"{ccy} {money(val)}" for ccy, val in sorted(payload["financing_sums"].get(period, {}).items())),
            }
        )
    write_sheet(ws, summary_rows, ["期间", "记录数", "缺成本记录数", "卖出总金额合计", "交易费用合计", "盈亏合计", "股息/分派合计", "股息/分派按年合计", "融资利息合计"])
    for period in payload["period_keys"]:
        ws = wb.create_sheet(period)
        period_rows = insert_order_summary(payload["rows"].get(period, []), headers)
        period_rows = append_period_totals(period_rows, headers, payload["income_sums"].get(period, {}), payload["financing_sums"].get(period, {}))
        write_sheet(ws, period_rows, headers)
    write_sheet(wb.create_sheet("股息利息_公司行动"), payload["income_items"], ["date", "category", "currency", "amount", "code", "name", "description", "source_file", "page", "market"])
    write_sheet(wb.create_sheet("融资利息"), payload["financing_items"], ["date", "category", "currency", "amount", "code", "name", "description", "source_file", "page", "market"])
    write_sheet(wb.create_sheet("缺成本与异常"), payload["exceptions"], sorted_headers(payload["exceptions"]))
    write_sheet(wb.create_sheet("缺名称与异常"), payload["name_exceptions"], sorted_headers(payload["name_exceptions"]))
    write_sheet(wb.create_sheet("解析来源与校验"), payload["file_summaries"], sorted_headers(payload["file_summaries"]))
    write_sheet(wb.create_sheet("交易来源明细"), payload["raw_trades"], sorted_headers(payload["raw_trades"]))
    write_sheet(wb.create_sheet("证券名称字典"), payload["security_master"], sorted_headers(payload["security_master"]))
    wb.save(path)


def sorted_headers(rows: list[dict[str, Any]]) -> list[str]:
    if not rows:
        return ["说明"]
    keys: list[str] = []
    for row in rows:
        for key in row:
            if key not in keys:
                keys.append(key)
    return keys


def output_name(market: str, regime: str, method: str) -> str:
    market_name = "港股" if market == "HK" else "美股"
    regime_name = "中国自然年" if regime == "calendar" else "香港财年"
    method_name = "FIFO" if method == "fifo" else "期间加权平均成本法"
    return f"{market_name}_{regime_name}_{method_name}_税务底稿.xlsx"


def build_payloads() -> list[dict[str, Any]]:
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=False)
    WORK_DIR.mkdir(parents=True, exist_ok=True)
    raw_trades, income, financing, file_summaries, parse_errors = parse_sources()
    security_master, name_exceptions = backfill_names(raw_trades)
    trades = normalized_trades(raw_trades)
    (WORK_DIR / "normalized_trades_audit.json").write_text(json.dumps(trades, ensure_ascii=False, indent=2), encoding="utf-8")
    (WORK_DIR / "source_parse_audit.json").write_text(json.dumps({"files": file_summaries, "parse_errors": parse_errors}, ensure_ascii=False, indent=2), encoding="utf-8")
    payloads = []
    for market in ["HK", "US"]:
        market_trades = [t for t in trades if t["market"] == market]
        market_income = [x for x in income if x.get("market") == market]
        market_financing = [x for x in financing if x.get("market") == market]
        for regime in ["calendar", "hk_fiscal"]:
            income_sums = income_period_sums(market_income, regime, market)
            financing_sums = income_period_sums(market_financing, regime, market)
            for method in ["fifo", "period_weighted_average"]:
                rows, exceptions = calculate_fifo(market_trades, regime) if method == "fifo" else calculate_period_weighted(market_trades, regime)
                fx_exceptions = add_fx_to_rows(rows, market)
                exceptions.extend(fx_exceptions)
                payloads.append(
                    {
                        "market": market,
                        "regime": regime,
                        "method": method,
                        "period_keys": period_keys(regime),
                        "rows": {k: rows.get(k, []) for k in period_keys(regime)},
                        "income_sums": {k: dict(v) for k, v in income_sums.items()},
                        "financing_sums": {k: dict(v) for k, v in financing_sums.items()},
                        "income_items": market_income,
                        "financing_items": market_financing,
                        "exceptions": exceptions,
                        "name_exceptions": [x for x in name_exceptions if x.get("market") == market],
                        "file_summaries": file_summaries,
                        "raw_trades": market_trades,
                        "security_master": [x for x in security_master if x.get("market") == market],
                    }
                )
    return payloads


def write_audit_only() -> dict[str, Any]:
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=False)
    WORK_DIR.mkdir(parents=True, exist_ok=True)
    raw_trades, income, financing, file_summaries, parse_errors = parse_sources()
    security_master, name_exceptions = backfill_names(raw_trades)
    trades = normalized_trades(raw_trades)
    audit = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_root": str(ROOT),
        "output_root": str(OUTPUT_ROOT),
        "records": {
            "raw_trades": len(raw_trades),
            "normalized_trades": len(trades),
            "income_items": len(income),
            "financing_items": len(financing),
            "parse_errors": len(parse_errors),
            "name_exceptions": len(name_exceptions),
        },
        "files": file_summaries,
        "parse_errors": parse_errors,
        "name_exceptions": name_exceptions,
    }
    (WORK_DIR / "normalized_trades_audit.json").write_text(json.dumps(trades, ensure_ascii=False, indent=2), encoding="utf-8")
    (WORK_DIR / "income_audit.json").write_text(json.dumps(income, ensure_ascii=False, indent=2), encoding="utf-8")
    (WORK_DIR / "financing_audit.json").write_text(json.dumps(financing, ensure_ascii=False, indent=2), encoding="utf-8")
    (WORK_DIR / "source_parse_audit.json").write_text(json.dumps({"files": file_summaries, "parse_errors": parse_errors}, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUTPUT_ROOT / "run_report.json").write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    (WORK_DIR / "run_report.json").write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    return audit


def run_workpaper(audit_only: bool = False) -> dict[str, Any]:
    if audit_only:
        report = write_audit_only()
        print(json.dumps(report["records"], ensure_ascii=False, indent=2), flush=True)
        return report
    payloads = build_payloads()
    outputs = []
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    for payload in payloads:
        path = OUTPUT_ROOT / output_name(payload["market"], payload["regime"], payload["method"])
        actual_path = path
        try:
            write_workbook(path, payload)
        except PermissionError:
            actual_path = path.with_name(f"{path.stem}_{stamp}{path.suffix}")
            write_workbook(actual_path, payload)
            print(f"LOCKED {path}; WROTE {actual_path}", flush=True)
        outputs.append(str(actual_path))
        print(f"WROTE {actual_path}", flush=True)
    report = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_root": str(ROOT),
        "output_root": str(OUTPUT_ROOT),
        "outputs": outputs,
        "summary": [
            {
                "file": output_name(p["market"], p["regime"], p["method"]),
                "period_rows": {k: len(v) for k, v in p["rows"].items()},
                "income_items": len(p["income_items"]),
                "financing_items": len(p["financing_items"]),
                "exceptions": len(p["exceptions"]),
                "name_exceptions": len(p["name_exceptions"]),
            }
            for p in payloads
        ],
    }
    (OUTPUT_ROOT / "run_report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    (WORK_DIR / "run_report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report["summary"], ensure_ascii=False, indent=2), flush=True)
    return report


def main() -> None:
    run_workpaper(audit_only=False)


if __name__ == "__main__":
    main()
