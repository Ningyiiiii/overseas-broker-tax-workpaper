"""Build workbooks matching the reference format.

Each workbook contains:
- 年度合计 (9 cols)
- One sheet per period (CY2021, FY2021-2022, etc.) — 18 cols, SELL rows only
- 股息利息_公司行动 (detail rows plus one period total row)
- 融资利息 (detail rows plus one period total row)
- 缺成本与异常 (16 cols, full row data)
- 缺名称与异常 (4 cols)
- 解析来源与校验 (5 cols)
- 交易来源明细 (19 cols)
- 证券名称字典 (3 cols)

Header fill is yellow (FFFF00), bold, centered.
Numeric fields are real Excel numbers.
"""

from __future__ import annotations

from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from tax_workpaper.engines.fifo import TradeRow
from tax_workpaper.engines.fx import format_fx_note, get_period_end_fx
from tax_workpaper.engines.periods import parse_date, period_end_date, period_keys_for
from tax_workpaper.normalize.schema import FinancingInterestRecord, IncomeRecord
from tax_workpaper.output.workbook_schema import MARKETS, output_filename

HEADER_FILL = PatternFill("solid", fgColor="FFFF00")
HEADER_FONT = Font(bold=True)


def _num(value):
    if value is None:
        return None
    if isinstance(value, str):
        try:
            return float(value.replace(",", ""))
        except ValueError:
            return value
    return value


def _write_header(ws, headers: list[str]) -> None:
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _set_widths(ws, widths: list[float]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Period sheet (18 cols, SELL only) ──────────────────────────────────────────

PERIOD_HEADERS = [
    "交易日期", "股票代码", "股票名称", "卖出价格", "买入价格", "交易数量",
    "卖出总金额", "买入总金额", "交易费用", "盈亏", "备注：买入时间", "币种",
    "订单编号", "来源", "异常", "股息/分派合计", "股息/分派按年合计", "融资利息合计",
]
PERIOD_WIDTHS = [12, 8, 10, 10, 10, 8, 12, 11, 9, 12, 15, 8, 11, 45, 30, 9, 11, 11]


def _build_period_sheet(
    ws,
    period_regime: str,
    period_key: str,
    trade_rows: list[TradeRow],
    income_rows: list[IncomeRecord],
    financing_rows: list[FinancingInterestRecord],
) -> None:
    period_index = 0 if period_regime == "china_calendar_year" else 1
    relevant = [
        r for r in trade_rows
        if (r.period_keys[period_index] if r.period_keys else "") == period_key
    ]
    period_income = [i for i in income_rows if _in_period(i.date, period_regime, period_key)]
    period_fin = [f for f in financing_rows if _in_period(f.date, period_regime, period_key)]

    _write_header(ws, PERIOD_HEADERS)
    _set_widths(ws, PERIOD_WIDTHS)

    row_idx = 2
    for r in relevant:
        if r.side != "SELL":
            continue
        if r.is_summary:
            cells = [
                r.trade_date, r.code, r.name, r.price, None, abs(r.quantity),
                _num(r.gross_amount), None, _num(r.fee_total), None,
                "订单汇总（不参与盈亏计算）", r.currency, r.raw_text.split("\n")[0][:20] if r.raw_text else "",
                r.source_note or "订单分拆成交汇总", "", "", "", "",
            ]
        else:
            buy_price = None
            if r.buy_allocated_amount is not None and abs(r.quantity) > 0:
                buy_price = r.buy_allocated_amount / abs(r.quantity)
            note = r.pwa_note or r.buy_trade_date or ""
            if r.missing_cost:
                note = "缺买入成本"
            order_id = ""
            if r.raw_text:
                m_id = r.raw_text.split("\n")[0]
                for part in m_id.split():
                    if part.isdigit() and len(part) >= 6:
                        order_id = part
                        break
            exception_note = "缺买入成本" if r.missing_cost else ""
            cells = [
                r.trade_date, r.code, r.name, r.price, _num(buy_price), abs(r.quantity),
                _num(r.gross_amount), _num(r.buy_allocated_amount), _num(r.transaction_fee),
                _num(r.pnl), note, r.currency, order_id,
                r.source_note or "", exception_note, "", "", "",
            ]
        for col, value in enumerate(cells, start=1):
            ws.cell(row=row_idx, column=col, value=value)
        row_idx += 1

    # Bottom total rows
    total_sell = sum(r.gross_amount for r in relevant if r.side == "SELL" and not r.is_summary)
    total_fee = sum(r.fee_total for r in relevant if r.side == "SELL" and not r.is_summary)
    total_pnl = sum((r.pnl or 0.0) for r in relevant if r.side == "SELL" and not r.is_summary)
    total_div = sum((i.amount or 0.0) for i in period_income if i.category in ("股息/分派", "股息/分派相关费用"))
    total_div_pos = sum((i.amount or 0.0) for i in period_income if i.category == "股息/分派" and (i.amount or 0.0) > 0)
    total_fin = sum((f.amount or 0.0) for f in period_fin)

    # 年度合计 row
    annual_cells = ["年度合计", "", "", "", "", "", _num(total_sell), "", _num(total_fee), _num(total_pnl),
                    "", "", "", "", "", "", "", ""]
    for col, value in enumerate(annual_cells, start=1):
        ws.cell(row=row_idx, column=col, value=value)
    row_idx += 1

    # 股息/分派合计 row
    if period_income:
        div_currency = next((i.currency for i in period_income if i.currency), "")
        div_cells = ["股息/分派合计", "", "", "", "", "", "", "", "", "", "", div_currency, "", "", "",
                     _num(total_div), _num(total_div_pos), ""]
        for col, value in enumerate(div_cells, start=1):
            ws.cell(row=row_idx, column=col, value=value)
        row_idx += 1

    # 融资利息合计 row
    if period_fin:
        fin_currency = next((f.currency for f in period_fin if f.currency), "")
        fin_cells = ["融资利息合计", "", "", "", "", "", "", "", "", "", "", fin_currency, "", "", "",
                     "", "", _num(total_fin)]
        for col, value in enumerate(fin_cells, start=1):
            ws.cell(row=row_idx, column=col, value=value)


def _in_period(record_date: str, period_regime: str, period_key: str) -> bool:
    d = parse_date(record_date)
    if d is None:
        return False
    return period_keys_for(d).get(period_regime, "") == period_key


# ── 年度合计 sheet (9 cols) ────────────────────────────────────────────────────

ANNUAL_HEADERS = [
    "期间", "记录数", "缺成本记录数", "卖出总金额合计", "交易费用合计", "盈亏合计",
    "股息/分派合计", "股息/分派按年合计", "融资利息合计",
]
ANNUAL_WIDTHS = [13, 8, 8, 13, 11, 12, 14, 14, 15]


def _summarize_period(period_key, period_regime, trade_rows, income_rows, financing_rows) -> dict:
    period_index = 0 if period_regime == "china_calendar_year" else 1
    relevant = [r for r in trade_rows if (r.period_keys[period_index] if r.period_keys else "") == period_key]
    sells = [r for r in relevant if r.side == "SELL" and not r.is_summary]
    period_income = [i for i in income_rows if _in_period(i.date, period_regime, period_key)]
    period_fin = [f for f in financing_rows if _in_period(f.date, period_regime, period_key)]
    total_div = sum((i.amount or 0.0) for i in period_income if i.category in ("股息/分派", "股息/分派相关费用"))
    total_div_pos = sum((i.amount or 0.0) for i in period_income if i.category == "股息/分派" and (i.amount or 0.0) > 0)
    return {
        "期间": period_key,
        "记录数": len(sells),
        "缺成本记录数": sum(1 for r in sells if r.missing_cost),
        "卖出总金额合计": _num(sum(r.gross_amount for r in sells)),
        "交易费用合计": _num(sum(r.fee_total for r in sells)),
        "盈亏合计": _num(sum((r.pnl or 0.0) for r in sells)),
        "股息/分派合计": f"HKD {total_div:.2f}" if period_income else "",
        "股息/分派按年合计": f"HKD {total_div_pos:.2f}" if period_income else "",
        "融资利息合计": f"HKD {sum((f.amount or 0.0) for f in period_fin):.2f}" if period_fin else "",
    }


def _build_annual_sheet(ws, summaries: list[dict]) -> None:
    _write_header(ws, ANNUAL_HEADERS)
    _set_widths(ws, ANNUAL_WIDTHS)
    for i, s in enumerate(summaries, start=2):
        for col, key in enumerate(ANNUAL_HEADERS, start=1):
            ws.cell(row=i, column=col, value=s.get(key, ""))


# ── 股息利息_公司行动 sheet (10 cols, flat) ─────────────────────────────────────

DIV_HEADERS = ["date", "category", "currency", "amount", "code", "name", "description", "source_file", "page", "market", "股息/分派按年合计"]
DIV_WIDTHS = [12, 11, 10, 11, 8, 8, 45, 45, 8, 8, 16]


def _period_key_for_date(date_text: str, period_regime: str) -> str:
    d = parse_date(date_text)
    if d is None:
        return ""
    return period_keys_for(d)[period_regime]


def _build_dividend_sheet(ws, income_rows: list[IncomeRecord], period_regime: str) -> None:
    _write_header(ws, DIV_HEADERS)
    _set_widths(ws, DIV_WIDTHS)
    row_idx = 2
    grouped: dict[tuple[str, str], list[IncomeRecord]] = defaultdict(list)
    for row in income_rows:
        grouped[(_period_key_for_date(row.date, period_regime), row.currency or "")].append(row)
    for (period_key, currency), rows in sorted(grouped.items(), key=lambda item: item[0]):
        for r in sorted(rows, key=lambda item: item.date):
            cells = [
                r.date, r.category, r.currency, _num(r.amount), r.code, r.name,
                (r.raw_text or "")[:200], Path(r.source_file).name if r.source_file else "",
                r.source_page or "", r.market, None,
            ]
            for col, value in enumerate(cells, start=1):
                ws.cell(row=row_idx, column=col, value=value)
            row_idx += 1
        total = sum((r.amount or 0.0) for r in rows)
        cells = [
            period_key, "年度股息/分派合计", currency, None, "", "",
            "", "", "", "", _num(total),
        ]
        for col, value in enumerate(cells, start=1):
            ws.cell(row=row_idx, column=col, value=value)
        row_idx += 1


# ── 融资利息 sheet (10 cols, flat) ──────────────────────────────────────────────

FIN_HEADERS = ["date", "category", "currency", "amount", "code", "name", "description", "source_file", "page", "market", "融资利息按年合计"]
FIN_WIDTHS = [12, 10, 10, 10, 8, 8, 45, 45, 8, 8, 16]


def _build_financing_sheet(ws, financing_rows: list[FinancingInterestRecord], period_regime: str) -> None:
    _write_header(ws, FIN_HEADERS)
    _set_widths(ws, FIN_WIDTHS)
    row_idx = 2
    grouped: dict[tuple[str, str], list[FinancingInterestRecord]] = defaultdict(list)
    for row in financing_rows:
        grouped[(_period_key_for_date(row.date, period_regime), row.currency or "")].append(row)
    for (period_key, currency), rows in sorted(grouped.items(), key=lambda item: item[0]):
        for r in sorted(rows, key=lambda item: item.date):
            cells = [
                r.date, "融资利息", r.currency, _num(r.amount), "", "",
                (r.raw_text or "")[:200], Path(r.source_file).name if r.source_file else "",
                r.source_page or "", r.market, None,
            ]
            for col, value in enumerate(cells, start=1):
                ws.cell(row=row_idx, column=col, value=value)
            row_idx += 1
        total = sum((r.amount or 0.0) for r in rows)
        cells = [
            period_key, "年度融资利息合计", currency, None, "", "",
            "", "", "", "", _num(total),
        ]
        for col, value in enumerate(cells, start=1):
            ws.cell(row=row_idx, column=col, value=value)
        row_idx += 1


# ── 缺成本与异常 sheet (16 cols, full row data) ─────────────────────────────────

COST_EXC_HEADERS = [
    "type", "交易日期", "股票代码", "股票名称", "卖出价格", "买入价格", "交易数量",
    "卖出总金额", "买入总金额", "交易费用", "盈亏", "备注：买入时间", "币种",
    "订单编号", "来源", "异常",
]
COST_EXC_WIDTHS = [20, 12, 8, 9, 10, 8, 8, 10, 8, 8, 8, 9, 8, 45, 45, 8]


def _build_cost_exception_sheet(ws, trade_rows: list[TradeRow]) -> None:
    _write_header(ws, COST_EXC_HEADERS)
    _set_widths(ws, COST_EXC_WIDTHS)
    row_idx = 2
    for r in trade_rows:
        if not r.missing_cost:
            continue
        order_id = ""
        if r.raw_text:
            for part in r.raw_text.split("\n")[0].split():
                if part.isdigit() and len(part) >= 6:
                    order_id = part
                    break
        cells = [
            "missing cost basis", r.trade_date, r.code, r.name, r.price, None,
            abs(r.quantity), _num(r.gross_amount), None, _num(r.fee_total), None,
            "缺买入成本", r.currency, order_id, r.source_file, "缺买入成本",
        ]
        for col, value in enumerate(cells, start=1):
            ws.cell(row=row_idx, column=col, value=value)
        row_idx += 1


# ── 缺名称与异常 sheet (4 cols) ─────────────────────────────────────────────────

NAME_EXC_HEADERS = ["market", "code", "exception", "source_file"]
NAME_EXC_WIDTHS = [8, 10, 16, 45]


def _build_name_exception_sheet(ws, name_exceptions: list[dict]) -> None:
    _write_header(ws, NAME_EXC_HEADERS)
    _set_widths(ws, NAME_EXC_WIDTHS)
    for i, exc in enumerate(name_exceptions, start=2):
        cells = [
            exc.get("market", ""), exc.get("code", ""),
            "缺股票名称或来源名称疑似乱码", exc.get("source_file", ""),
        ]
        for col, value in enumerate(cells, start=1):
            ws.cell(row=i, column=col, value=value)


# ── 解析来源与校验 sheet (5 cols) ───────────────────────────────────────────────

SRC_HEADERS = ["file", "trades", "income_items", "financing_items", "error"]
SRC_WIDTHS = [45, 8, 14, 17, 8]


def _build_source_sheet(ws, source_files: list[dict]) -> None:
    _write_header(ws, SRC_HEADERS)
    _set_widths(ws, SRC_WIDTHS)
    for i, src in enumerate(source_files, start=2):
        cells = [
            src.get("文件", ""), src.get("交易笔数", 0), src.get("股息/分派笔数", 0),
            src.get("融资利息笔数", 0), src.get("异常", ""),
        ]
        for col, value in enumerate(cells, start=1):
            ws.cell(row=i, column=col, value=value)


# ── 交易来源明细 sheet (19 cols) ─────────────────────────────────────────────────

TRADE_DETAIL_HEADERS = [
    "side", "code", "name", "market", "currency", "trade_datetime", "trade_date",
    "settle_date", "quantity", "price", "amount", "fee_total", "change_amount",
    "order_id", "fill_id", "source_file", "source_pages", "format", "notes",
]
TRADE_DETAIL_WIDTHS = [8, 8, 12, 8, 10, 21, 12, 13, 10, 11, 11, 11, 15, 45, 45, 45, 14, 26, 45]


def _build_trade_detail_sheet(ws, trade_rows: list[TradeRow]) -> None:
    _write_header(ws, TRADE_DETAIL_HEADERS)
    _set_widths(ws, TRADE_DETAIL_WIDTHS)
    for i, r in enumerate(trade_rows, start=2):
        order_id = ""
        if r.raw_text:
            for part in r.raw_text.split("\n")[0].split():
                if part.isdigit() and len(part) >= 6:
                    order_id = part
                    break
        side = "sell" if r.side == "SELL" else "buy"
        change = r.gross_amount if r.side == "SELL" else -(r.gross_amount + r.fee_total)
        cells = [
            side, r.code, r.name, r.market, r.currency, r.trade_date, r.trade_date,
            "", abs(r.quantity), r.price, _num(r.gross_amount), _num(r.fee_total),
            _num(change), order_id, order_id,
            Path(r.source_file).name if r.source_file else "", r.source_page or "",
            "", "",  # broker 列（不再硬编码，留空）
        ]
        for col, value in enumerate(cells, start=1):
            ws.cell(row=i, column=col, value=value)


# ── 证券名称字典 sheet (3 cols) ──────────────────────────────────────────────────

SM_HEADERS = ["market", "code", "name"]
SM_WIDTHS = [8, 8, 39]


def _build_security_master_sheet(ws, master: dict[tuple[str, str], str]) -> None:
    _write_header(ws, SM_HEADERS)
    _set_widths(ws, SM_WIDTHS)
    for i, ((market, code), name) in enumerate(sorted(master.items()), start=2):
        ws.cell(row=i, column=1, value=market)
        ws.cell(row=i, column=2, value=code)
        ws.cell(row=i, column=3, value=name)


# ── Main builder ────────────────────────────────────────────────────────────────

def build_workbook(
    output_path: Path,
    market: str,
    period_regime: str,
    cost_method: str,
    trade_rows: list[TradeRow],
    income_rows: list[IncomeRecord],
    financing_rows: list[FinancingInterestRecord],
    source_files: list[dict],
    master: dict[tuple[str, str], str],
    exceptions: list[dict],
    name_exceptions: list[dict],
    period_label: str,
) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    period_index = 0 if period_regime == "china_calendar_year" else 1
    period_keys_set: set[str] = set()
    for r in trade_rows:
        if r.period_keys and r.period_keys[period_index]:
            period_keys_set.add(r.period_keys[period_index])
    for r in income_rows:
        d = parse_date(r.date)
        if d:
            period_keys_set.add(period_keys_for(d)[period_regime])
    for f in financing_rows:
        d = parse_date(f.date)
        if d:
            period_keys_set.add(period_keys_for(d)[period_regime])
    period_keys_sorted = sorted(period_keys_set)
    if not period_keys_sorted:
        period_keys_sorted = [f"CY{date.today().year}" if period_regime == "china_calendar_year"
                              else f"FY{date.today().year - 1}-{date.today().year}"]

    summaries = [
        _summarize_period(pk, period_regime, trade_rows, income_rows, financing_rows)
        for pk in period_keys_sorted
    ]

    # 年度合计
    ws = wb.create_sheet("年度合计")
    _build_annual_sheet(ws, summaries)

    # Per-period sheets
    for pk in period_keys_sorted:
        ws = wb.create_sheet(pk)
        _build_period_sheet(ws, period_regime, pk, trade_rows, income_rows, financing_rows)
        if market == "US":
            pe = period_end_date(period_regime, pk)
            fx = get_period_end_fx("USDHKD", pe.isoformat())
            note = format_fx_note("USDHKD", fx)
            ws.cell(row=ws.max_row + 2, column=1, value="汇率说明")
            ws.cell(row=ws.max_row, column=2, value=note)

    # 股息利息_公司行动
    ws = wb.create_sheet("股息利息_公司行动")
    _build_dividend_sheet(ws, income_rows, period_regime)

    # 融资利息
    ws = wb.create_sheet("融资利息")
    _build_financing_sheet(ws, financing_rows, period_regime)

    # 缺成本与异常
    ws = wb.create_sheet("缺成本与异常")
    _build_cost_exception_sheet(ws, trade_rows)

    # 缺名称与异常
    ws = wb.create_sheet("缺名称与异常")
    _build_name_exception_sheet(ws, name_exceptions)

    # 解析来源与校验
    ws = wb.create_sheet("解析来源与校验")
    _build_source_sheet(ws, source_files)

    # 交易来源明细
    ws = wb.create_sheet("交易来源明细")
    _build_trade_detail_sheet(ws, trade_rows)

    # 证券名称字典
    ws = wb.create_sheet("证券名称字典")
    _build_security_master_sheet(ws, master)

    wb.save(output_path)
    return output_path
