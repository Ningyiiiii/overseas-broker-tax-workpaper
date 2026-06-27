"""Redacted regression fixtures for the public skill runtime.

Run directly:
    python tests/test_regression_fixtures.py

These tests avoid real brokerage statements. They lock the behavior that tends
to regress when the skill is generalized for more brokers and statement layouts.
"""

from __future__ import annotations

import importlib.util
import shutil
import sys
import tempfile
import types
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"


def bootstrap_tax_workpaper() -> None:
    if str(SCRIPTS) not in sys.path:
        sys.path.insert(0, str(SCRIPTS))
    pkg = types.ModuleType("tax_workpaper")
    pkg.__path__ = [str(SCRIPTS)]
    sys.modules["tax_workpaper"] = pkg


bootstrap_tax_workpaper()

from tax_workpaper.engines.fifo import calculate_fifo
from tax_workpaper.engines.period_weighted_average import calculate_period_weighted_average
from tax_workpaper.normalize.schema import FinancingInterestRecord, IncomeRecord, TradeRecord
from tax_workpaper.output.build_workbook import build_workbook
from tax_workpaper.parsers.futu_parser import FutuParser, classify_market, infer_side, is_real_trade_candidate


def load_run_workpaper_module():
    spec = importlib.util.spec_from_file_location("run_workpaper_under_test", SCRIPTS / "run_workpaper.py")
    module = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(module)
    return module


def trade(
    side: str,
    date: str,
    code: str,
    qty: float,
    price: float,
    gross: float,
    fee: float,
    market: str = "HK",
    currency: str = "HKD",
    source_row: int = 1,
) -> TradeRecord:
    return TradeRecord(
        broker="fixture",
        market=market,
        currency=currency,
        code=code,
        name="测试股份",
        side=side,
        trade_date=date,
        settle_date=None,
        order_id=f"{code}-{date}-{side}",
        trade_id=f"{code}-{date}-{side}-{source_row}",
        quantity=qty,
        price=price,
        gross_amount=gross,
        fee_total=fee,
        source_file="fixture_statement.pdf",
        source_page=1,
        source_row=source_row,
        raw_text="redacted fixture row",
    )


def test_entrypoint_is_portable_and_registers_futu() -> None:
    module = load_run_workpaper_module()
    assert module._SKILL_ROOT == ROOT
    assert "trae-cn" not in str(module._SKILL_ROOT).lower()
    assert "futu" in [parser.broker for parser in module._load_parsers()]


def test_config_password_candidates_support_common_shapes() -> None:
    module = load_run_workpaper_module()
    tmp = Path(tempfile.mkdtemp(prefix="obtw_config_fixture_"))
    try:
        (tmp / "config").mkdir()
        (tmp / "config" / "config.json").write_text(
            '{"markets":["HK"],"password_file":"passwords.json","passwords":["111"],"per_broker":{"futu":["222"]}}',
            encoding="utf-8",
        )
        (tmp / "config" / "passwords.json").write_text(
            '{"password_candidates":["333"],"per_account":{"100123":"444"},"per_file":{"a.pdf":"555"}}',
            encoding="utf-8",
        )
        args = types.SimpleNamespace(config="", password=["000"])
        config, passwords = module._load_config_and_passwords(args, tmp)
        assert config["markets"] == ["HK"]
        assert passwords[:6] == ["000", "111", "222", "333", "444", "555"]
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_audit_only_outputs_normalized_records_without_full_calculation() -> None:
    module = load_run_workpaper_module()
    tmp = Path(tempfile.mkdtemp(prefix="obtw_audit_fixture_"))
    try:
        output_dir = tmp / "out"
        output_dir.mkdir()
        records = [
            trade("BUY", "2024-01-02", "00363", 1000, 1.0, 1000.0, 10.0),
            trade("SELL", "2024-06-14", "00363", -500, 1.5, 750.0, 8.0, source_row=2),
        ]
        report = module.write_audit_only(
            output_dir=output_dir,
            sources=[tmp / "fixture_statement.pdf"],
            results=[{"broker": "fixture"}],
            errors=[],
            all_trades=records,
            all_income=[],
            all_financing=[],
            validations=[],
            name_excs=[],
            source_files_meta=[{"文件": "fixture_statement.pdf", "交易笔数": 2, "股息/分派笔数": 0}],
            broker_stats={"fixture": 1},
        )
        assert Path(report["audit_workbook"]).exists()
        assert (output_dir / "run_report.json").exists()
        assert not list(output_dir.glob("*_税务底稿.xlsx"))
        wb = load_workbook(report["audit_workbook"], data_only=True)
        assert {"扫描文件", "解析来源与校验", "交易", "异常"}.issubset(set(wb.sheetnames))
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_futu_parser_detection_and_trade_heuristics() -> None:
    parser = FutuParser()
    assert parser.can_parse_with_text("Futu Securities Account No. 1001231828219038 SEHK")
    assert classify_market("SEHK", "HKD").market == "HK"
    assert classify_market("EDGX", "USD").market == "US"
    assert classify_market("SEHK", "USD").market == ""
    side, note = infer_side("buy open", "100.00")
    assert side == ""
    assert "conflict" in note
    base = {
        "exchange": "SEHK",
        "currency": "HKD",
        "trade_date": "2025/10/02",
        "trade_time": "09:46:07",
        "settle_date": "2025/10/06",
        "quantity": "8000",
        "price": "4.59",
        "gross_amount": "36720.00",
        "cash_change": "36652.32",
        "raw_text": "sell close normal execution row",
    }
    assert is_real_trade_candidate(base)
    assert not is_real_trade_candidate({**base, "raw_text": "Account Upgrade cash movement"})


def test_fifo_split_and_missing_cost_are_explicit() -> None:
    records = [
        trade("BUY", "2024-01-02", "00363", 1000, 10.0, 10000.0, 10.0),
        trade("BUY", "2024-02-02", "00363", 500, 12.0, 6000.0, 6.0, source_row=2),
        trade("SELL", "2024-03-03", "00363", -1800, 15.0, 27000.0, 18.0, source_row=3),
    ]
    result = calculate_fifo(records, "HK")
    sell_details = [row for row in result.rows if row.side == "SELL" and row.is_split_detail]
    missing = [row for row in sell_details if row.missing_cost]
    assert len(sell_details) == 3
    assert [abs(row.quantity) for row in sell_details] == [1000, 500, 300]
    assert len(missing) == 1
    assert missing[0].buy_allocated_amount is None
    assert missing[0].pnl is None
    assert any(exc["type"] == "missing_cost" for exc in result.exceptions)


def test_period_weighted_average_uses_same_period_later_buys() -> None:
    records = [
        trade("SELL", "2024-01-10", "00363", -100, 20.0, 2000.0, 20.0, source_row=1),
        trade("BUY", "2024-12-20", "00363", 100, 10.0, 1000.0, 10.0, source_row=2),
    ]
    result = calculate_period_weighted_average(records, "HK", "china_calendar_year", opening_positions={})
    sell_rows = [row for row in result.rows if row.side == "SELL"]
    assert len(sell_rows) == 1
    assert sell_rows[0].buy_allocated_amount == 1010.0
    assert sell_rows[0].pnl == 970.0


def test_workbook_adds_period_totals_for_dividends_and_financing() -> None:
    tmp = Path(tempfile.mkdtemp(prefix="obtw_workbook_fixture_"))
    try:
        target = tmp / "fixture.xlsx"
        income = [
            IncomeRecord("fixture", "HK", "HKD", "2024-06-14", "00363", "上海实业", "股息/分派", 3120.0, None, None, "fixture.pdf", 1, None, "F/D-HKD"),
            IncomeRecord("fixture", "HK", "HKD", "2024-06-14", "00363", "上海实业", "股息/分派相关费用", -39.0, None, None, "fixture.pdf", 1, None, "Scrip Charge"),
        ]
        financing = [
            FinancingInterestRecord("fixture", "HK", "HKD", "2024-07-01", -12.5, "fixture.pdf", 1, None, "Margin interest"),
        ]
        build_workbook(
            target,
            market="HK",
            period_regime="china_calendar_year",
            cost_method="fifo",
            trade_rows=[],
            income_rows=income,
            financing_rows=financing,
            source_files=[],
            master={},
            exceptions=[],
            name_exceptions=[],
            period_label="中国自然年",
        )
        wb = load_workbook(target, data_only=True)
        div_rows = list(wb["股息利息_公司行动"].iter_rows(values_only=True))
        fin_rows = list(wb["融资利息"].iter_rows(values_only=True))
        assert div_rows[-1][0] == "CY2024"
        assert div_rows[-1][1] == "年度股息/分派合计"
        assert div_rows[-1][-1] == 3081.0
        assert fin_rows[-1][0] == "CY2024"
        assert fin_rows[-1][1] == "年度融资利息合计"
        assert fin_rows[-1][-1] == -12.5
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def main() -> int:
    tests = [
        test_entrypoint_is_portable_and_registers_futu,
        test_config_password_candidates_support_common_shapes,
        test_audit_only_outputs_normalized_records_without_full_calculation,
        test_futu_parser_detection_and_trade_heuristics,
        test_fifo_split_and_missing_cost_are_explicit,
        test_period_weighted_average_uses_same_period_later_buys,
        test_workbook_adds_period_totals_for_dividends_and_financing,
    ]
    for test in tests:
        test()
        print(f"PASS {test.__name__}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
